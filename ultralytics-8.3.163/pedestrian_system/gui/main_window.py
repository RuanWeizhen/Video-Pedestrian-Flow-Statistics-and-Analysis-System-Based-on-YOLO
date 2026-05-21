from pathlib import Path
from datetime import datetime
from collections import deque
import time
import csv
import os
import platform
import shutil

import sys
import subprocess
import shlex
import yaml
from PyQt5.QtCore import Qt, QThread, QTimer, pyqtSignal
from PyQt5.QtGui import QBrush, QColor, QFont, QIcon, QPainter, QPen, QPixmap
from PyQt5.QtWidgets import (
    QAction,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QDoubleSpinBox,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QLineEdit,
    QInputDialog,
    QMessageBox,
    QPushButton,
    QFrame,
    QSplitter,
    QScrollArea,
    QStackedWidget,
    QTabWidget,
    QSpinBox,
    QListWidget,
    QListWidgetItem,
    QSystemTrayIcon,
    QVBoxLayout,
    QWidget,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QStyle,
)

from openpyxl import Workbook

from .annotation_panel import AnnotationPanel
from .control_panel import ControlPanel
from .event_table_panel import EventTablePanel
from .log_panel import LogPanel
from .stats_panel import StatsPanel
from .system_info_panel import SystemInfoPanel
from .trend_panel import TrendChartWidget, TotalTrendWidget, FpsChartWidget
from .video_panel import VideoPanel
from .video_player import VideoPlayerWidget
from .dashboard_widgets import DashboardMetricCard, DashboardSection, SidebarButton, UserBadge, build_metric_row
from .experiment_table_panel import ExperimentTablePanel
from .trajectory_replay_dialog import TrajectoryReplayDialog
from .worker import WorkerThread
from utils.auth_manager import AuthManager
from .employee import EmployeeManagementPage
from utils.db_manager import DatabaseManager
from utils.model_registry import (
    activate_model,
    ensure_model_record,
    get_active_model,
    get_model_by_id,
    get_model_by_path,
    infer_framework,
    load_registry,
    normalize_registry,
    remove_model,
    save_registry,
    update_model_evaluation,
)
from utils.coordinate_transform import frame_points_to_frame_points
from utils.paths import resource_path, writable_path
from utils.torch_runtime import ensure_torch_preloaded


class ToolOptionsDialog(QDialog):
    def __init__(self, title: str, fields: list[tuple[str, QWidget]], parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self._fields = {}

        for label, widget in fields:
            self._fields[label] = widget
            form.addRow(label, widget)

        layout.addLayout(form)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def value(self, label: str):
        widget = self._fields[label]
        if isinstance(widget, QSpinBox):
            return int(widget.value())
        if isinstance(widget, QDoubleSpinBox):
            return float(widget.value())
        if isinstance(widget, QCheckBox):
            return bool(widget.isChecked())
        if isinstance(widget, QComboBox):
            data = widget.currentData()
            return data if data is not None else widget.currentText()
        if isinstance(widget, QLineEdit):
            return widget.text().strip()
        return None


class ToolRunnerThread(QThread):
    log = pyqtSignal(str)
    finished = pyqtSignal(int)

    def __init__(self, cmd: list[str], cwd: str | None = None):
        super().__init__()
        self.cmd = cmd
        self.cwd = cwd

    def run(self):
        try:
            self.log.emit(f"运行命令: {' '.join(self.cmd)}")
            proc = subprocess.Popen(self.cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, cwd=self.cwd)
            assert proc.stdout is not None
            for line in proc.stdout:
                self.log.emit(line.rstrip())
            proc.wait()
            self.finished.emit(proc.returncode)
        except Exception as exc:
            self.log.emit(f"工具运行异常: {exc}")
            self.finished.emit(-1)


class InlineExportThread(QThread):
    log = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, weights: str, out_dir: str, imgsz: int, opset: int,
                 simplify: bool, dynamic: bool, half: bool):
        super().__init__()
        self.weights = weights
        self.out_dir = out_dir
        self.imgsz = imgsz
        self.opset = opset
        self.simplify = simplify
        self.dynamic = dynamic
        self.half = half

    def run(self):
        try:
            import shutil
            from pathlib import Path

            weights_path = Path(self.weights).resolve()
            output_dir = Path(self.out_dir).resolve()
            output_dir.mkdir(parents=True, exist_ok=True)

            self.log.emit(f"=== ONNX 导出开始 ===")
            self.log.emit(f"权重文件: {weights_path}")
            self.log.emit(f"imgsz={self.imgsz} opset={self.opset} simplify={self.simplify} half={self.half}")

            from ultralytics import YOLO
            self.log.emit(f"加载模型: {weights_path}")
            model = YOLO(str(weights_path))
            self.log.emit("模型加载完成，开始导出...")

            exported_path = model.export(
                format="onnx",
                imgsz=self.imgsz,
                opset=self.opset,
                simplify=False,
                dynamic=self.dynamic,
                half=self.half,
                project=str(output_dir),
                name=weights_path.stem,
            )

            exported_path = Path(str(exported_path)).resolve()
            self.log.emit(f"导出路径: {exported_path}")

            target_path = output_dir / f"{weights_path.stem}.onnx"

            if not exported_path.exists():
                onnx_files = list(output_dir.rglob("*.onnx"))
                if onnx_files:
                    exported_path = onnx_files[0].resolve()
                    self.log.emit(f"搜索到: {exported_path}")
                else:
                    self.log.emit("ERROR: 未找到任何 .onnx 文件")
                    self.finished.emit(False, "")
                    return

            if exported_path != target_path:
                self.log.emit(f"扁平化: {exported_path} -> {target_path}")
                shutil.copy2(exported_path, target_path)

            if self.simplify:
                self.log.emit("正在运行 onnxsim 简化...")
                try:
                    from onnxsim import simplify
                    import onnx
                    original_size = target_path.stat().st_size / (1024 * 1024)
                    model_onnx = onnx.load(str(target_path))
                    model_simplified, check = simplify(model_onnx)
                    if not check:
                        self.log.emit("WARNING: onnxsim 简化校验未通过，保留未简化版本")
                    else:
                        onnx.save(model_simplified, str(target_path))
                        simplified_size = target_path.stat().st_size / (1024 * 1024)
                        self.log.emit(f"简化完成: {original_size:.2f} MB -> {simplified_size:.2f} MB")
                except ImportError:
                    self.log.emit("WARNING: onnxsim 未安装，跳过简化步骤。请运行: pip install onnxsim")
                except Exception as sim_exc:
                    self.log.emit(f"WARNING: onnxsim 简化失败，保留原始导出: {sim_exc}")

            file_size_mb = target_path.stat().st_size / (1024 * 1024)
            self.log.emit(f"SUCCESS: {target_path} ({file_size_mb:.2f} MB)")
            self.finished.emit(True, str(target_path))

        except Exception as exc:
            import traceback
            self.log.emit(f"ERROR: {type(exc).__name__}: {exc}")
            self.log.emit(traceback.format_exc())
            self.finished.emit(False, "")


class ModelEvaluationThread(QThread):
    log = pyqtSignal(str)
    done = pyqtSignal(dict)
    failed = pyqtSignal(str)

    def __init__(
        self,
        model_path: str,
        data_path: str,
        imgsz: int,
        batch: int,
        device,
        split: str = "val",
        augment: bool = False,
        parent=None,
    ):
        super().__init__(parent)
        self.model_path = str(model_path)
        self.data_path = str(data_path)
        self.imgsz = int(imgsz)
        self.batch = int(batch)
        self.device = device
        self.split = str(split or "val")
        self.augment = bool(augment)

    def run(self):
        try:
            from ultralytics.models.yolo.model import YOLO

            start_time = time.perf_counter()
            self.log.emit(f"开始评估模型: {self.model_path}")
            self.log.emit(f"数据集: {self.data_path} | imgsz={self.imgsz} | batch={self.batch} | device={self.device}")
            model = YOLO(self.model_path)
            metrics = model.val(
                data=self.data_path,
                imgsz=self.imgsz,
                batch=self.batch,
                device=self.device,
                split=self.split,
                augment=self.augment,
            )
            elapsed_seconds = time.perf_counter() - start_time

            results_dict = {}
            try:
                results_dict = dict(metrics.results_dict)
            except Exception:
                pass

            box = getattr(metrics, "box", None)
            precision = float(results_dict.get("metrics/precision(B)", getattr(box, "p", 0.0) if box is not None else 0.0))
            recall = float(results_dict.get("metrics/recall(B)", getattr(box, "r", 0.0) if box is not None else 0.0))
            map50 = float(results_dict.get("metrics/mAP50(B)", getattr(box, "map50", 0.0) if box is not None else 0.0))
            map5095 = float(results_dict.get("metrics/mAP50-95(B)", getattr(box, "map", 0.0) if box is not None else 0.0))
            fitness = float(results_dict.get("fitness", getattr(metrics, "fitness", 0.0)))

            speed = getattr(metrics, "speed", {}) or {}
            inference_ms = float(speed.get("inference", 0.0) or 0.0)
            fps = (1000.0 / inference_ms) if inference_ms > 1e-9 else 0.0

            payload = {
                "model_path": self.model_path,
                "data_path": self.data_path,
                "imgsz": self.imgsz,
                "batch": self.batch,
                "device": self.device,
                "split": self.split,
                "augment": self.augment,
                "precision": precision,
                "recall": recall,
                "map50": map50,
                "map5095": map5095,
                "fitness": fitness,
                "fps": fps,
                "inference_ms": inference_ms,
                "elapsed_seconds": elapsed_seconds,
            }
            self.done.emit(payload)
        except Exception as exc:
            self.failed.emit(str(exc))


class MainWindow(QMainWindow):
    def __init__(self, current_user: dict | None = None):
        super().__init__()
        self.current_user = current_user or {"username": "未登录", "role": "管理员"}
        self.setWindowTitle("视频行人客流量统计与分析系统")
        self.resize(1200, 800)

        self.current_config_path = None
        self.current_config = self._default_config()
        self.model_registry = load_registry(self.current_config.get("detector", {}).get("model_path", ""))
        self._model_eval_thread = None
        self.export_thread = None
        self.auth_manager = AuthManager()
        self.history_db = DatabaseManager(writable_path("outputs/traffic.db"))
        self.tool_runners = []
        self.recent_events = deque(maxlen=8)
        self.current_source_fps = 25.0
        self.current_video_resolution = "-"
        self.current_source_frame_size = None
        self.current_runtime_start = None
        self.current_frame_number = 0
        self.current_total_frames = 0
        self.current_total_people = 0
        self.current_up_count = 0
        self.current_down_count = 0
        self.current_detecting = False
        self.current_source_name = "未选择"
        self.current_avg_fps = 0.0
        self.current_device = self._device_label()
        self.current_detector_type = self._detector_type_label()
        self.current_tracker_type = "DeepSORT"
        self.current_onnx_enabled = False
        self.current_session_id = None
        self.current_run_id = None
        self.event_history = []
        self.trend_history = []
        self.fps_history = []
        self._active_alert_keys = set()
        self._alert_last_emit_at = {}
        self._last_realtime_stats = None

        self.worker = WorkerThread()
        self.worker.set_config(self.current_config)
        self.init_ui()
        self._init_system_tray()
        self._init_menus()
        self.connect_signals()
        self.apply_runtime_params(self.annotation_panel.get_params())
        self.apply_user_permissions()

    def init_ui(self):
        self.setObjectName("MainWindow")

        self.video_panel = VideoPanel()
        self.video_player = VideoPlayerWidget(self.video_panel)
        self.log_panel = LogPanel()
        self.event_table_panel = EventTablePanel()
        self.filtered_event_table_panel = EventTablePanel()
        self.experiment_table_panel = ExperimentTablePanel()
        self.system_info_panel = SystemInfoPanel()
        self.control_panel = ControlPanel()
        self.stats_panel = StatsPanel()
        self.annotation_panel = AnnotationPanel()

        # 检测页仅保留实时检测相关按钮，模型/配置类入口迁移到系统管理页
        self.control_panel.btn_export_onnx.hide()
        self.control_panel.btn_quantize.hide()
        self.control_panel.btn_benchmark.hide()

        self.home_trend_panel = TotalTrendWidget()
        self.stats_history_panel = TrendChartWidget()
        self.stats_fps_panel = FpsChartWidget()

        self.home_metric_total = DashboardMetricCard("总通行人数", "0", "本次会话累计")
        self.home_metric_up = DashboardMetricCard("Up", "0", "上行人数")
        self.home_metric_down = DashboardMetricCard("Down", "0", "下行人数")
        self.home_metric_current = DashboardMetricCard("当前人数", "0", "当前画面")
        self.home_metric_fps = DashboardMetricCard("FPS", "0.0", "实时性能")

        self.detect_metric_total = DashboardMetricCard("总通行人数", "0", "实时统计")
        self.detect_metric_up = DashboardMetricCard("Up", "0", "上行人数")
        self.detect_metric_down = DashboardMetricCard("Down", "0", "下行人数")
        self.detect_metric_current = DashboardMetricCard("当前人数", "0", "当前画面")
        self.detect_metric_fps = DashboardMetricCard("FPS", "0.0", "实时性能")

        self.sidebar_buttons: dict[str, SidebarButton] = {}
        self.page_titles: dict[str, tuple[str, str]] = {
            "home": ("系统首页", "查看当前运行状态、关键指标与简化趋势"),
            "detection": ("行人检测", "视频画面、运行控制和参数标注集中管理"),
            "management": ("系统管理", "模型管理、配置管理和系统维护"),
            "statistics": ("数据统计", "历史分析、参数实验和结果导出"),
            "profile": ("个人信息", "查看个人资料和账号安全信息"),
            "password": ("员工管理", "管理员工账号、角色和账号状态"),
        }

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        root_layout = QHBoxLayout(main_widget)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        self.sidebar_widget = self._build_sidebar()
        self.content_widget = QWidget()
        content_layout = QVBoxLayout(self.content_widget)
        content_layout.setContentsMargins(18, 18, 18, 18)
        content_layout.setSpacing(14)

        self.header_widget = self._build_header()
        content_layout.addWidget(self.header_widget)

        self.pages = QStackedWidget()
        content_layout.addWidget(self.pages, 1)

        root_layout.addWidget(self.sidebar_widget)
        root_layout.addWidget(self.content_widget, 1)

        self._build_pages()
        self._apply_dashboard_style()
        self._switch_page("home")

    def _build_sidebar(self):
        sidebar = QWidget()
        sidebar.setObjectName("Sidebar")
        sidebar.setFixedWidth(220)
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(16, 18, 16, 16)
        layout.setSpacing(10)

        brand = QWidget()
        brand_layout = QVBoxLayout(brand)
        brand_layout.setContentsMargins(16, 20, 16, 20)
        brand_layout.setSpacing(6)
        brand_title = QLabel("客流统计系统")
        brand_title.setObjectName("SidebarBrandTitle")
        brand_subtitle = QLabel("Dashboard UI")
        brand_subtitle.setObjectName("SidebarBrandSubtitle")
        brand_layout.addWidget(brand_title)
        brand_layout.addWidget(brand_subtitle)
        layout.addWidget(brand)

        for key, text in (
            ("home", "系统首页"),
            ("detection", "行人检测"),
            ("management", "系统管理"),
            ("statistics", "数据统计"),
            ("profile", "个人中心"),
            ("password", "员工管理"),
        ):
            button = SidebarButton(text)
            button.clicked.connect(lambda _=False, page_key=key: self._switch_page(page_key))
            self.sidebar_buttons[key] = button
            layout.addWidget(button)

        layout.addStretch()

        self.btn_logout = QPushButton("退出登录")
        self.btn_logout.setObjectName("SidebarLogoutButton")
        self.btn_logout.setCursor(Qt.PointingHandCursor)
        self.btn_logout.clicked.connect(self.close)
        layout.addWidget(self.btn_logout)
        return sidebar
        layout.addWidget(self.btn_logout)
        return sidebar

    def _build_header(self):
        header = QWidget()
        header.setObjectName("PageHeader")
        layout = QHBoxLayout(header)
        layout.setContentsMargins(6, 4, 6, 4)
        layout.setSpacing(12)

        title_box = QWidget()
        title_layout = QVBoxLayout(title_box)
        title_layout.setContentsMargins(0, 0, 0, 0)
        title_layout.setSpacing(2)
        self.page_title_label = QLabel("系统首页")
        self.page_title_label.setObjectName("PageTitle")
        self.page_desc_label = QLabel("总览关键指标、趋势和最近运行摘要")
        self.page_desc_label.setObjectName("PageSubtitle")
        title_layout.addWidget(self.page_title_label)
        title_layout.addWidget(self.page_desc_label)

        self.user_badge = UserBadge("当前用户", self.current_user.get("username", "未登录"))
        self.role_badge = UserBadge("身份", self.current_user.get("role", "管理员"))

        layout.addWidget(title_box, 1)
        layout.addWidget(self.user_badge)
        layout.addWidget(self.role_badge)
        return header

    def _make_scroll_page(self, widget: QWidget):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setWidget(widget)
        return scroll

    def _build_pages(self):
        self.page_widgets = {}

        home_page = self._build_home_page()
        detection_page = self._build_detection_page()
        management_page = self._build_management_page()
        statistics_page = self._build_statistics_page()
        profile_page = self._build_profile_page()
        password_page = self._build_password_page()

        for key, widget in (
            ("home", home_page),
            ("detection", detection_page),
            ("management", management_page),
            ("statistics", statistics_page),
            ("profile", profile_page),
            ("password", password_page),
        ):
            self.page_widgets[key] = widget
            # 统计页内部已经实现了独立的 QScrollArea，避免嵌套滚动条
            if key == "statistics":
                self.pages.addWidget(widget)
            else:
                self.pages.addWidget(self._make_scroll_page(widget))

    def _build_home_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(14)

        metrics_row = build_metric_row([
            self.home_metric_total,
            self.home_metric_up,
            self.home_metric_down,
            self.home_metric_current,
            self.home_metric_fps,
        ])
        layout.addLayout(metrics_row)

        info_section = DashboardSection("运行概览", "当前检测状态、运行上下文与最近事件")
        info_body = QWidget()
        info_layout = QHBoxLayout(info_body)
        info_layout.setContentsMargins(0, 0, 0, 0)
        info_layout.setSpacing(24)

        status_form = QFormLayout()
        status_form.setSpacing(4)
        self.home_status_detect = QLabel("未运行")
        self.home_status_source = QLabel("未选择")
        self.home_status_model = QLabel("未加载")
        self.home_status_config = QLabel("未加载")
        for lbl in (self.home_status_detect, self.home_status_source,
                     self.home_status_model, self.home_status_config):
            lbl.setWordWrap(True)
            lbl.setMinimumWidth(120)
        status_form.addRow("检测状态", self.home_status_detect)
        status_form.addRow("视频源", self.home_status_source)
        status_form.addRow("模型状态", self.home_status_model)
        status_form.addRow("配置状态", self.home_status_config)
        info_layout.addLayout(status_form)

        summary_layout = QVBoxLayout()
        summary_layout.setSpacing(4)
        summary_title = QLabel("最近事件")
        summary_title.setStyleSheet("font-weight: 600; color: #16324f;")
        self.home_summary_list = QListWidget()
        self.home_summary_list.setMaximumHeight(120)
        summary_layout.addWidget(summary_title)
        summary_layout.addWidget(self.home_summary_list, 1)
        info_layout.addLayout(summary_layout, 1)

        info_section.set_body_widget(info_body)
        layout.addWidget(info_section)

        trend_section = DashboardSection("累计通行人数趋势", "实时累计通行人数（单调递增，数据中断时保持最后值）")
        self.home_trend_panel.setMinimumHeight(380)
        trend_section.set_body_widget(self.home_trend_panel)
        layout.addWidget(trend_section, 1)

        self._refresh_home_status_card()

        return page

    def _build_detection_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(14)

        metrics_row = build_metric_row([
            self.detect_metric_total,
            self.detect_metric_up,
            self.detect_metric_down,
            self.detect_metric_current,
            self.detect_metric_fps,
        ])
        layout.addLayout(metrics_row)

        splitter = QSplitter(Qt.Horizontal)

        video_section = DashboardSection("视频画面", "显示检测画面与标注交互")
        self.video_placeholder = QLabel("请选择并开始视频")
        self.video_placeholder.setAlignment(Qt.AlignCenter)
        self.video_placeholder.setStyleSheet(
            "color: rgba(255,255,255,0.6); font-size: 18px; font-weight: 600;"
            "background-color: #1a1a2e; border-radius: 12px;"
        )
        self.video_placeholder.setMinimumHeight(300)
        video_section.body_layout.addWidget(self.video_player, 1)
        video_section.body_layout.addWidget(self.video_placeholder)
        self.video_placeholder.hide()

        right_tabs = QTabWidget()
        right_tabs.setMaximumWidth(520)

        control_tab = QWidget()
        control_layout = QVBoxLayout(control_tab)
        control_layout.setContentsMargins(0, 0, 0, 0)
        control_layout.setSpacing(8)
        control_layout.addWidget(self.control_panel)

        self.alert_toggle_btn = QPushButton("▶ 实时告警设置")
        self.alert_toggle_btn.setCheckable(True)
        self.alert_toggle_btn.setChecked(False)
        self.alert_toggle_btn.setStyleSheet(
            "QPushButton { text-align: left; padding: 8px 14px; font-weight: 600;"
            "border: 1px solid #dbe6f2; border-radius: 10px; background: #f8fafc; }"
            "QPushButton:hover { background: #eef3fb; }"
            "QPushButton:checked { background: #e8f0fe; color: #2f80ed; }"
        )
        self.alert_toggle_btn.toggled.connect(self._toggle_alert_section)
        control_layout.addWidget(self.alert_toggle_btn)

        self.realtime_alert_section = self._build_realtime_alert_section()
        self.realtime_alert_section.setVisible(False)
        control_layout.addWidget(self.realtime_alert_section)
        control_layout.addStretch()
        right_tabs.addTab(control_tab, "运行控制")

        annotation_tab = QWidget()
        annotation_layout = QVBoxLayout(annotation_tab)
        annotation_layout.setContentsMargins(0, 0, 0, 0)
        annotation_layout.addWidget(self.annotation_panel)
        annotation_layout.addStretch()
        right_tabs.addTab(annotation_tab, "参数与标注")

        info_tab = QWidget()
        info_layout = QVBoxLayout(info_tab)
        info_layout.setContentsMargins(0, 0, 0, 0)
        info_tabs = QTabWidget()
        log_tab = QWidget()
        log_layout = QVBoxLayout(log_tab)
        log_layout.setContentsMargins(0, 0, 0, 0)
        log_layout.addWidget(self.log_panel)
        info_tabs.addTab(log_tab, "运行日志")

        sys_tab = QWidget()
        sys_layout = QVBoxLayout(sys_tab)
        sys_layout.setContentsMargins(0, 0, 0, 0)
        sys_layout.addWidget(self.system_info_panel)
        info_tabs.addTab(sys_tab, "系统信息")
        info_layout.addWidget(info_tabs)
        right_tabs.addTab(info_tab, "状态详情")

        splitter.addWidget(video_section)
        splitter.addWidget(right_tabs)
        splitter.setSizes([820, 360])
        splitter.setStretchFactor(0, 7)
        splitter.setStretchFactor(1, 3)
        layout.addWidget(splitter, 1)
        return page

    def _build_statistics_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(14)

        query_section = DashboardSection("查询条件", "按时间、视频、模式和批次筛选统计数据")
        query_widget = QWidget()
        query_layout = QHBoxLayout(query_widget)
        query_layout.setContentsMargins(0, 0, 0, 0)
        query_layout.setSpacing(10)

        self.query_time_start = QLineEdit("")
        self.query_time_start.setPlaceholderText("YYYY-MM-DD HH:MM:SS")
        self.query_time_end = QLineEdit("")
        self.query_time_end.setPlaceholderText("YYYY-MM-DD HH:MM:SS")
        self.query_stats_mode = QComboBox()
        self.query_stats_mode.addItems(["最新一次检测", "指定检测批次", "历史累计统计"])
        self.query_video_name = QComboBox()
        self.query_video_name.addItem("全部视频")
        self.query_run_batch = QComboBox()
        self.query_run_batch.setMinimumWidth(260)
        self.query_direction = QComboBox()
        self.query_direction.addItems(["All", "Up", "Down"])
        self.query_bucket = QComboBox()
        self.query_bucket.addItems(["按分钟", "每5分钟"])
        self.btn_query = QPushButton("查询")
        self.btn_query_reset = QPushButton("重置")
        self.btn_query.clicked.connect(self.apply_statistics_filters)
        self.btn_query_reset.clicked.connect(self.reset_statistics_filters)
        self.query_stats_mode.currentIndexChanged.connect(self._refresh_statistics_run_choices)
        self.query_video_name.currentIndexChanged.connect(self._refresh_statistics_run_choices)
        self.query_run_batch.currentIndexChanged.connect(self.apply_statistics_filters)

        query_layout.addWidget(QLabel("模式"))
        query_layout.addWidget(self.query_stats_mode)
        query_layout.addWidget(QLabel("开始"))
        query_layout.addWidget(self.query_time_start)
        query_layout.addWidget(QLabel("结束"))
        query_layout.addWidget(self.query_time_end)
        query_layout.addWidget(QLabel("视频"))
        query_layout.addWidget(self.query_video_name)
        query_layout.addWidget(QLabel("批次"))
        query_layout.addWidget(self.query_run_batch)
        query_layout.addWidget(QLabel("方向"))
        query_layout.addWidget(self.query_direction)
        query_layout.addWidget(QLabel("统计粒度"))
        query_layout.addWidget(self.query_bucket)
        query_layout.addWidget(self.btn_query)
        query_layout.addWidget(self.btn_query_reset)
        query_section.set_body_widget(query_widget)
        layout.addWidget(query_section)

        # 把剩余内容放入可滚动区域：保证左侧侧栏和顶部 header 固定
        scroll_container = QWidget()
        scroll_layout = QVBoxLayout(scroll_container)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(14)

        summary_row = build_metric_row([
            DashboardMetricCard("本次检测总数", "0", "最新一次检测"),
            DashboardMetricCard("本次 Up", "0", "最新一次检测"),
            DashboardMetricCard("本次 Down", "0", "最新一次检测"),
            DashboardMetricCard("平均 FPS", "0.0", "最新一次检测"),
        ])
        summary_widget = QWidget()
        summary_widget.setLayout(summary_row)
        summary_widget.setMinimumHeight(100)
        summary_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.stats_query_total = summary_row.itemAt(0).widget()
        self.stats_query_up = summary_row.itemAt(1).widget()
        self.stats_query_down = summary_row.itemAt(2).widget()
        self.stats_query_fps = summary_row.itemAt(3).widget()
        self.stats_query_section = query_section
        scroll_layout.addWidget(summary_widget)

        # 趋势图与 FPS: 提高高度，确保图表不会被压缩
        chart_splitter = QSplitter(Qt.Horizontal)
        history_section = DashboardSection("历史趋势图", "展示当前模式下的 Up/Down/Total 历史变化")
        history_section.set_body_widget(self.stats_history_panel)
        history_section.setMinimumHeight(440)
        self.stats_history_panel.setMinimumHeight(420)
        self.stats_history_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        fps_section = DashboardSection("FPS 曲线", "展示当前批次的 FPS 采样变化")
        fps_section.set_body_widget(self.stats_fps_panel)
        fps_section.setMinimumHeight(360)
        self.stats_fps_panel.setMinimumHeight(320)
        self.stats_fps_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.stats_history_section = history_section
        self.stats_fps_section = fps_section

        chart_splitter.addWidget(history_section)
        chart_splitter.addWidget(fps_section)
        chart_splitter.setSizes([760, 420])
        scroll_layout.addWidget(chart_splitter)

        # 事件表
        table_section = DashboardSection("事件表", "当前模式下的事件与批次 ID")
        table_section.set_body_widget(self.filtered_event_table_panel)
        table_section.setMinimumHeight(460)
        self.filtered_event_table_panel.setMinimumHeight(420)
        self.filtered_event_table_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.stats_event_section = table_section
        scroll_layout.addWidget(table_section)

        # 参数实验表
        exp_section = DashboardSection("参数实验结果表", "当前模式下的会话性能与计数结果")
        exp_section.set_body_widget(self.experiment_table_panel)
        exp_section.setMinimumHeight(340)
        self.experiment_table_panel.setMinimumHeight(300)
        self.experiment_table_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.stats_experiment_section = exp_section
        scroll_layout.addWidget(exp_section)

        # 导出区域放在底部，可滚动查看
        export_section = DashboardSection("导出", "导出当前模式下的事件表、统计结果和图表截图")
        export_widget = QWidget()
        # 使用 FlowLayout 或多行布局处理按钮过多问题
        export_layout = QHBoxLayout(export_widget)
        export_layout.setContentsMargins(16, 16, 16, 16)
        export_layout.setSpacing(16)
        
        self.btn_export_event_table = QPushButton("导出事件表")
        self.btn_export_stats_result = QPushButton("导出统计结果")
        self.btn_export_chart_snapshot = QPushButton("导出图表截图")
        self.btn_export_history_data = QPushButton("导出数据库历史结果")
        self.btn_open_replay = QPushButton("轨迹回放")
        
        for btn in (self.btn_export_event_table, self.btn_export_stats_result, 
                self.btn_export_chart_snapshot, self.btn_export_history_data, self.btn_open_replay):
            btn.setMinimumHeight(44)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            export_layout.addWidget(btn)

        self.btn_export_event_table.clicked.connect(self.export_filtered_events)
        self.btn_export_stats_result.clicked.connect(self.on_export_results_requested)
        self.btn_export_chart_snapshot.clicked.connect(self.export_statistics_charts)
        self.btn_export_history_data.clicked.connect(self.export_database_history_results)
        self.btn_open_replay.clicked.connect(self.open_selected_batch_replay)
        
        export_section.set_body_widget(export_widget)
        # 不要写死小高度，自适应内容
        export_section.setMinimumHeight(140)
        export_section.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Minimum)
        self.stats_export_section = export_section
        scroll_layout.addWidget(export_section)

        # 添加底部留白以避免按钮贴底
        bottom_spacer = QWidget()
        bottom_spacer.setMinimumHeight(20)
        scroll_layout.addWidget(bottom_spacer)

        # 总体滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setWidget(scroll_container)
        layout.addWidget(scroll, 1)

        self.reset_statistics_filters()

        return page

    def _build_management_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setSpacing(14)

        scroll_container = QWidget()
        scroll_layout = QVBoxLayout(scroll_container)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(14)

        top_row = QHBoxLayout()

        model_card = DashboardSection("模型管理", "当前模型信息与模型工具")
        model_body = QWidget()
        model_layout = QVBoxLayout(model_body)
        model_layout.setContentsMargins(0, 0, 0, 0)
        self.management_model_path_label = QLabel("-")
        self.management_model_name_label = QLabel("-")
        self.management_model_type_label = QLabel("-")
        self.management_model_version_label = QLabel("-")
        self.management_model_status_label = QLabel("-")
        self.management_model_count_label = QLabel("-")
        self.management_model_eval_label = QLabel("-")
        self.management_imgsz_label = QLabel("-")
        self.management_device_label = QLabel("-")
        for label in (
            self.management_model_path_label,
            self.management_model_name_label,
            self.management_model_type_label,
            self.management_model_version_label,
            self.management_model_status_label,
            self.management_model_count_label,
            self.management_model_eval_label,
            self.management_imgsz_label,
            self.management_device_label,
        ):
            label.setWordWrap(True)
        model_form = QFormLayout()
        model_form.addRow("当前模型路径", self.management_model_path_label)
        model_form.addRow("模型名称", self.management_model_name_label)
        model_form.addRow("模型类型", self.management_model_type_label)
        model_form.addRow("模型版本", self.management_model_version_label)
        model_form.addRow("模型状态", self.management_model_status_label)
        model_form.addRow("注册模型数", self.management_model_count_label)
        model_form.addRow("最近评估", self.management_model_eval_label)
        model_form.addRow("输入尺寸", self.management_imgsz_label)
        model_form.addRow("当前设备", self.management_device_label)
        model_layout.addLayout(model_form)

        model_btn_row1 = QHBoxLayout()
        self.btn_export_onnx_page = QPushButton("导出 ONNX")
        self.btn_quantize_page = QPushButton("量化 ONNX")
        self.btn_benchmark_page = QPushButton("Benchmark ONNX")
        self.btn_export_onnx_page.clicked.connect(self.on_export_onnx_requested)
        self.btn_quantize_page.clicked.connect(self.on_quantize_onnx_requested)
        self.btn_benchmark_page.clicked.connect(self.on_benchmark_onnx_requested)
        for button in (self.btn_export_onnx_page, self.btn_quantize_page, self.btn_benchmark_page):
            model_btn_row1.addWidget(button)
        model_btn_row1.addStretch()
        model_layout.addLayout(model_btn_row1)

        model_card.set_body_widget(model_body)
        model_card.setMinimumHeight(250)
        top_row.addWidget(model_card)

        config_card = DashboardSection("配置管理", "当前配置文件与配置维护")
        config_body = QWidget()
        config_layout = QVBoxLayout(config_body)
        config_layout.setContentsMargins(0, 0, 0, 0)
        self.management_config_path_label = QLabel("-")
        self.management_config_summary_label = QLabel("-")
        self.management_config_summary_label.setWordWrap(True)
        config_form = QFormLayout()
        config_form.addRow("当前配置文件路径", self.management_config_path_label)
        config_form.addRow("查看配置摘要", self.management_config_summary_label)
        config_layout.addLayout(config_form)

        config_btn_row1 = QHBoxLayout()
        self.btn_import_config_page = QPushButton("导入配置")
        self.btn_save_config_page = QPushButton("保存配置")
        self.btn_restore_default_config = QPushButton("恢复默认配置")
        self.btn_view_config_summary = QPushButton("查看配置摘要")
        self.btn_import_config_page.clicked.connect(self.load_yaml_config)
        self.btn_save_config_page.clicked.connect(self.save_yaml_config)
        self.btn_restore_default_config.clicked.connect(self.restore_default_config)
        self.btn_view_config_summary.clicked.connect(self.show_config_summary)
        for button in (self.btn_import_config_page, self.btn_save_config_page, self.btn_restore_default_config, self.btn_view_config_summary):
            config_btn_row1.addWidget(button)
        config_btn_row1.addStretch()
        config_layout.addLayout(config_btn_row1)

        config_card.set_body_widget(config_body)
        config_card.setMinimumHeight(250)
        top_row.addWidget(config_card)

        scroll_layout.addLayout(top_row)

        middle_row = QHBoxLayout()

        runtime_card = DashboardSection("运行环境", "Python、PyTorch、OpenCV 和 CUDA 状态")
        runtime_body = QWidget()
        runtime_layout = QGridLayout(runtime_body)
        runtime_layout.setSpacing(6)
        runtime_layout.setContentsMargins(8, 8, 8, 8)
        self.management_python_label = QLabel("-")
        self.management_torch_label = QLabel("-")
        self.management_opencv_label = QLabel("-")
        self.management_cuda_label = QLabel("-")
        self.management_run_device_label = QLabel("-")
        self.management_platform_label = QLabel("-")
        for lbl in (self.management_python_label, self.management_torch_label,
                     self.management_opencv_label, self.management_cuda_label,
                     self.management_run_device_label, self.management_platform_label):
            lbl.setWordWrap(True)

        runtime_layout.addWidget(QLabel("Python"), 0, 0)
        runtime_layout.addWidget(self.management_python_label, 0, 1)
        runtime_layout.addWidget(QLabel("PyTorch"), 1, 0)
        runtime_layout.addWidget(self.management_torch_label, 1, 1)
        runtime_layout.addWidget(QLabel("OpenCV"), 2, 0)
        runtime_layout.addWidget(self.management_opencv_label, 2, 1)
        runtime_layout.addWidget(QLabel("CUDA"), 3, 0)
        runtime_layout.addWidget(self.management_cuda_label, 3, 1)
        runtime_layout.addWidget(QLabel("设备"), 4, 0)
        runtime_layout.addWidget(self.management_run_device_label, 4, 1)
        runtime_layout.addWidget(QLabel("系统"), 5, 0)
        runtime_layout.addWidget(self.management_platform_label, 5, 1)
        runtime_card.set_body_widget(runtime_body)
        runtime_card.setMinimumHeight(220)
        middle_row.addWidget(runtime_card)

        maintenance_card = DashboardSection("系统维护", "输出目录、日志目录和临时文件维护")
        maintenance_body = QWidget()
        maintenance_layout = QVBoxLayout(maintenance_body)
        maintenance_layout.setContentsMargins(0, 0, 0, 0)
        maintenance_layout.setSpacing(8)
        maint_btn_grid = QGridLayout()
        maint_btn_grid.setSpacing(8)
        self.btn_open_output_dir = QPushButton("打开输出目录")
        self.btn_open_log_dir = QPushButton("打开日志目录")
        self.btn_clear_cache = QPushButton("清理缓存")
        self.btn_clear_temp_files = QPushButton("清空临时文件")
        self.btn_view_system_log = QPushButton("查看系统日志")
        self.btn_open_output_dir.clicked.connect(self.open_output_dir)
        self.btn_open_log_dir.clicked.connect(self.open_log_dir)
        self.btn_clear_cache.clicked.connect(self.clear_cache)
        self.btn_clear_temp_files.clicked.connect(self.clear_temp_files)
        self.btn_view_system_log.clicked.connect(self.show_system_log)
        maint_btn_grid.addWidget(self.btn_open_output_dir, 0, 0)
        maint_btn_grid.addWidget(self.btn_open_log_dir, 0, 1)
        maint_btn_grid.addWidget(self.btn_clear_cache, 1, 0)
        maint_btn_grid.addWidget(self.btn_clear_temp_files, 1, 1)
        maint_btn_grid.addWidget(self.btn_view_system_log, 2, 0, 1, 2)
        maintenance_layout.addLayout(maint_btn_grid)
        self.management_maintenance_hint = QLabel("用于维护运行目录、临时文件与系统日志。")
        self.management_maintenance_hint.setWordWrap(True)
        maintenance_layout.addWidget(self.management_maintenance_hint)
        maintenance_card.set_body_widget(maintenance_body)
        maintenance_card.setMinimumHeight(240)
        middle_row.addWidget(maintenance_card)

        scroll_layout.addLayout(middle_row)

        registry_card = DashboardSection("模型版本管理", "添加模型版本、激活当前模型并同步评估记录")
        registry_body = QWidget()
        registry_layout = QVBoxLayout(registry_body)
        registry_layout.setContentsMargins(0, 0, 0, 0)
        registry_btn_row = QHBoxLayout()
        self.btn_add_model_page = QPushButton("添加模型")
        self.btn_activate_model_page = QPushButton("激活选中")
        self.btn_remove_model_page = QPushButton("删除选中")
        self.btn_evaluate_model_page = QPushButton("评估选中")
        self.btn_refresh_model_registry_page = QPushButton("刷新列表")
        self.btn_add_model_page.clicked.connect(self.add_model_version)
        self.btn_activate_model_page.clicked.connect(self.activate_selected_model)
        self.btn_remove_model_page.clicked.connect(self.remove_selected_model)
        self.btn_evaluate_model_page.clicked.connect(self.evaluate_selected_model)
        self.btn_refresh_model_registry_page.clicked.connect(self.refresh_model_registry_view)
        for button in (
            self.btn_add_model_page,
            self.btn_activate_model_page,
            self.btn_remove_model_page,
            self.btn_evaluate_model_page,
            self.btn_refresh_model_registry_page,
        ):
            button.setMinimumHeight(40)
            registry_btn_row.addWidget(button)
        registry_btn_row.addStretch()
        registry_layout.addLayout(registry_btn_row)
        self.management_model_registry_table = QTableWidget(0, 11)
        self.management_model_registry_table.setHorizontalHeaderLabels([
            "状态",
            "模型名称",
            "版本",
            "框架",
            "路径",
            "推理配置",
            "评估次数",
            "最近评估",
            "最近结果",
            "更新时间",
            "备注",
        ])
        self.management_model_registry_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.management_model_registry_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.management_model_registry_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.management_model_registry_table.setMinimumHeight(320)
        self.management_model_registry_table.verticalHeader().setDefaultSectionSize(36)
        self.management_model_registry_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.management_model_registry_table.horizontalHeader().setStretchLastSection(False)
        self.management_model_registry_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        registry_layout.addWidget(self.management_model_registry_table)
        self.management_registry_hint = QLabel("选中模型后可直接激活、删除或执行验证评估。")
        self.management_registry_hint.setWordWrap(True)
        registry_layout.addWidget(self.management_registry_hint)
        registry_card.set_body_widget(registry_body)
        registry_card.setMinimumHeight(360)
        scroll_layout.addWidget(registry_card)

        eval_card = DashboardSection("模型评估历史", "记录验证指标、推理速度和耗时")
        eval_body = QWidget()
        eval_layout = QVBoxLayout(eval_body)
        eval_layout.setContentsMargins(0, 0, 0, 0)
        self.management_model_eval_table = QTableWidget(0, 13)
        self.management_model_eval_table.setHorizontalHeaderLabels([
            "时间",
            "类型",
            "模型名称",
            "版本",
            "数据集",
            "设备",
            "Precision",
            "Recall",
            "mAP50",
            "mAP50-95",
            "FPS",
            "耗时",
            "备注",
        ])
        self.management_model_eval_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.management_model_eval_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.management_model_eval_table.setMinimumHeight(300)
        self.management_model_eval_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.management_model_eval_table.horizontalHeader().setStretchLastSection(False)
        self.management_model_eval_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        eval_layout.addWidget(self.management_model_eval_table)
        eval_card.set_body_widget(eval_body)
        eval_card.setMinimumHeight(340)
        scroll_layout.addWidget(eval_card)

        benchmark_card = DashboardSection("Benchmark 结果表", "展示 ONNX Benchmark 历史结果")
        benchmark_body = QWidget()
        benchmark_layout = QVBoxLayout(benchmark_body)
        benchmark_layout.setContentsMargins(0, 0, 0, 0)
        self.management_benchmark_table = QTableWidget(0, 7)
        self.management_benchmark_table.setHorizontalHeaderLabels(["测试时间", "模型名称", "输入尺寸", "平均 FPS", "推理耗时", "设备", "备注"])
        self.management_benchmark_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.management_benchmark_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.management_benchmark_table.setMinimumHeight(260)
        self.management_benchmark_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.management_benchmark_table.horizontalHeader().setStretchLastSection(False)
        self.management_benchmark_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        benchmark_layout.addWidget(self.management_benchmark_table)
        benchmark_card.set_body_widget(benchmark_body)
        benchmark_card.setMinimumHeight(300)
        scroll_layout.addWidget(benchmark_card)

        log_card = DashboardSection("操作日志", "最近系统管理操作")
        log_body = QWidget()
        log_layout = QVBoxLayout(log_body)
        log_layout.setContentsMargins(0, 0, 0, 0)
        self.management_log_table = QTableWidget(0, 5)
        self.management_log_table.setHorizontalHeaderLabels(["时间", "操作类型", "操作内容", "操作结果", "备注"])
        self.management_log_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.management_log_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.management_log_table.setMinimumHeight(220)
        self.management_log_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.management_log_table.horizontalHeader().setStretchLastSection(False)
        self.management_log_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        log_layout.addWidget(self.management_log_table)
        log_card.set_body_widget(log_body)
        log_card.setMinimumHeight(280)
        scroll_layout.addWidget(log_card)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setWidget(scroll_container)
        layout.addWidget(scroll, 1)

        self._sync_model_registry_from_current_config(save=True)
        self._load_sample_management_logs()
        self._refresh_management_page()
        self.refresh_model_registry_view()
        self.refresh_model_eval_table()
        self.refresh_management_benchmark_table()
        self.refresh_management_log_table()

        layout.addStretch()
        return page

    def _build_profile_page(self):
        page = QWidget()
        page.setObjectName("ProfilePage")
        page.setStyleSheet("QWidget#ProfilePage { background: #F5F8FC; }")

        outer_wrapper = QHBoxLayout(page)
        outer_wrapper.setContentsMargins(24, 0, 24, 0)

        content = QWidget()
        content.setMaximumWidth(1100)
        outer_layout = QVBoxLayout(content)
        outer_layout.setContentsMargins(0, 0, 0, 24)
        outer_layout.setSpacing(18)

        outer_layout.addWidget(self._build_profile_banner())
        outer_layout.addLayout(self._build_profile_info_cards())
        outer_layout.addWidget(self._build_activity_log_card())
        outer_layout.addStretch()

        outer_wrapper.addStretch()
        outer_wrapper.addWidget(content)
        outer_wrapper.addStretch()

        return page

    def _card_frame(self, obj_name="", padding=(20, 20, 20, 20)):
        card = QFrame()
        if obj_name:
            card.setObjectName(obj_name)
        card.setStyleSheet(
            "QFrame#{obj} {{"
            "  background: #ffffff;"
            "  border: 1px solid #EEF1F5;"
            "  border-radius: 12px;"
            "}}".format(obj=obj_name) if obj_name else
            "QFrame {"
            "  background: #ffffff;"
            "  border: 1px solid #EEF1F5;"
            "  border-radius: 12px;"
            "}"
        )
        layout = QVBoxLayout(card)
        layout.setContentsMargins(*padding)
        layout.setSpacing(12)
        return card, layout

    def _make_badge(self, text, bg_color, text_color):
        badge = QLabel(text)
        badge.setStyleSheet(
            "background: {bg}; color: {fg}; font-size: 11px; font-weight: 600;"
            "padding: 3px 10px; border-radius: 8px;".format(bg=bg_color, fg=text_color)
        )
        badge.setFixedHeight(22)
        badge.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
        return badge

    def _make_card_row(self, label_text, value_attr, value_default="-"):
        row = QHBoxLayout()
        row.setSpacing(8)
        lbl = QLabel(label_text)
        lbl.setStyleSheet("color: #9CA3AF; font-size: 12px; font-weight: 500;")
        val = QLabel(value_default)
        val.setStyleSheet("color: #1F2937; font-size: 13px; font-weight: 600;")
        val.setWordWrap(True)
        row.addWidget(lbl)
        row.addWidget(val, 1)
        if value_attr:
            setattr(self, value_attr, val)
        return row

    def _build_profile_banner(self):
        card = QFrame()
        card.setObjectName("ProfileBanner")
        card.setStyleSheet(
            "QFrame#ProfileBanner {"
            "  background: qlineargradient(x1:0, y1:0, x2:1, y2:0,"
            "    stop:0 #E8F2FF, stop:1 #FFFFFF);"
            "  border: 1px solid #D6E4F5;"
            "  border-radius: 12px;"
            "}"
        )
        banner_layout = QHBoxLayout(card)
        banner_layout.setContentsMargins(28, 24, 28, 24)
        banner_layout.setSpacing(20)

        self.profile_avatar = QLabel()
        self.profile_avatar.setFixedSize(72, 72)
        self.profile_avatar.setAlignment(Qt.AlignCenter)
        banner_layout.addWidget(self.profile_avatar)

        info_col = QVBoxLayout()
        info_col.setSpacing(6)

        name_row = QHBoxLayout()
        name_row.setSpacing(10)
        self.profile_banner_name = QLabel("-")
        self.profile_banner_name.setStyleSheet("font-size: 20px; font-weight: 700; color: #1F2937;")
        name_row.addWidget(self.profile_banner_name)

        self.profile_banner_role_badge = self._make_badge("管理员", "#DBEAFE", "#1E40AF")
        name_row.addWidget(self.profile_banner_role_badge)

        self.profile_banner_status_badge = self._make_badge("已启用", "#D1FAE5", "#065F46")
        name_row.addWidget(self.profile_banner_status_badge)
        name_row.addStretch()
        info_col.addLayout(name_row)

        self.profile_banner_desc = QLabel("系统管理员账号，可管理客流检测、数据统计与系统配置。")
        self.profile_banner_desc.setStyleSheet("color: #6B7280; font-size: 13px;")
        info_col.addWidget(self.profile_banner_desc)

        banner_layout.addLayout(info_col, 1)

        btn_col = QVBoxLayout()
        btn_col.setSpacing(8)

        self.btn_change_password_page = QPushButton("🔒 修改密码")
        self.btn_change_password_page.setCursor(Qt.PointingHandCursor)
        self.btn_change_password_page.setFixedWidth(130)
        self.btn_change_password_page.setStyleSheet(
            "QPushButton {"
            "  background: #2F80ED; color: #ffffff; font-size: 13px; font-weight: 600;"
            "  border: none; border-radius: 8px; padding: 10px 0px;"
            "}"
            "QPushButton:hover { background: #2563EB; }"
            "QPushButton:pressed { background: #1D4ED8; }"
        )
        self.btn_change_password_page.clicked.connect(self._open_change_password_dialog)
        btn_col.addWidget(self.btn_change_password_page)

        self.btn_logout_page = QPushButton("🚪 退出登录")
        self.btn_logout_page.setCursor(Qt.PointingHandCursor)
        self.btn_logout_page.setFixedWidth(130)
        self.btn_logout_page.setStyleSheet(
            "QPushButton {"
            "  background: #F3F4F6; color: #374151; font-size: 13px; font-weight: 600;"
            "  border: none; border-radius: 8px; padding: 10px 0px;"
            "}"
            "QPushButton:hover { background: #EF4444; color: #ffffff; }"
            "QPushButton:pressed { background: #DC2626; color: #ffffff; }"
        )
        self.btn_logout_page.clicked.connect(self.close)
        btn_col.addWidget(self.btn_logout_page)
        btn_col.addStretch()

        banner_layout.addLayout(btn_col)
        return card

    def _build_profile_info_cards(self):
        row_layout = QHBoxLayout()
        row_layout.setSpacing(18)

        basic_card, basic_layout = self._card_frame("BasicInfoCard")
        basic_title = QLabel("📋 基础信息")
        basic_title.setStyleSheet("font-size: 15px; font-weight: 700; color: #1F2937;")
        basic_layout.addWidget(basic_title)
        basic_layout.addLayout(self._make_card_row("用户名：", "profile_basic_username"))
        basic_layout.addLayout(self._make_card_row("角色：", "profile_basic_role"))
        basic_layout.addLayout(self._make_card_row("账号状态：", "profile_basic_status"))
        basic_layout.addLayout(self._make_card_row("注册时间：", "profile_basic_reg_time"))
        basic_layout.addStretch()

        login_card, login_layout = self._card_frame("LoginInfoCard")
        login_title = QLabel("📊 登录信息")
        login_title.setStyleSheet("font-size: 15px; font-weight: 700; color: #1F2937;")
        login_layout.addWidget(login_title)
        login_layout.addLayout(self._make_card_row("上次登录：", "profile_login_last"))
        login_layout.addLayout(self._make_card_row("登录次数：", "profile_login_count"))
        login_tip = QLabel("💡 建议定期修改密码以保证账户安全")
        login_tip.setStyleSheet("color: #9CA3AF; font-size: 12px;")
        login_layout.addWidget(login_tip)
        login_layout.addStretch()

        sec_card, sec_layout = self._card_frame("SecurityCard")
        sec_title = QLabel("🔐 账号安全")
        sec_title.setStyleSheet("font-size: 15px; font-weight: 700; color: #1F2937;")
        sec_layout.addWidget(sec_title)
        sec_layout.addLayout(self._make_card_row("密码状态：", "profile_sec_pwd_status", "正常"))
        sec_layout.addLayout(self._make_card_row("权限等级：", "profile_sec_perm_level"))
        sec_layout.addLayout(self._make_card_row("账号保护：", "profile_sec_protection", "已开启"))
        btn_change_sec = QPushButton("🔒 修改密码")
        btn_change_sec.setCursor(Qt.PointingHandCursor)
        btn_change_sec.setStyleSheet(
            "QPushButton {"
            "  background: #2F80ED; color: #ffffff; font-size: 12px; font-weight: 600;"
            "  border: none; border-radius: 6px; padding: 8px 16px;"
            "}"
            "QPushButton:hover { background: #2563EB; }"
        )
        btn_change_sec.clicked.connect(self._open_change_password_dialog)
        sec_btn_row = QHBoxLayout()
        sec_btn_row.addWidget(btn_change_sec)
        sec_btn_row.addStretch()
        sec_layout.addLayout(sec_btn_row)
        sec_layout.addStretch()

        row_layout.addWidget(basic_card)
        row_layout.addWidget(login_card)
        row_layout.addWidget(sec_card)

        return row_layout

    def _build_activity_log_card(self):
        card, card_layout = self._card_frame("ActivityLogCard", padding=(20, 20, 20, 20))

        header_row = QHBoxLayout()
        header = QLabel("📝 最近操作记录")
        header.setStyleSheet("font-size: 15px; font-weight: 700; color: #1F2937;")
        header_row.addWidget(header)
        header_row.addStretch()
        card_layout.addLayout(header_row)

        self.profile_activity_table = QTableWidget(0, 4)
        self.profile_activity_table.setHorizontalHeaderLabels(["时间", "操作", "状态", "说明"])
        self.profile_activity_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.profile_activity_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.profile_activity_table.horizontalHeader().setStretchLastSection(True)
        self.profile_activity_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.profile_activity_table.verticalHeader().setVisible(False)
        self.profile_activity_table.setFixedHeight(180)
        self.profile_activity_table.setStyleSheet(
            "QTableWidget { border: none; font-size: 13px; }"
            "QHeaderView::section { background: #F9FAFB; color: #6B7280; font-weight: 600;"
            "  border: none; border-bottom: 1px solid #EEF1F5; padding: 10px 8px; }"
        )
        card_layout.addWidget(self.profile_activity_table)

        self.profile_empty_state = QWidget()
        empty_layout = QVBoxLayout(self.profile_empty_state)
        empty_layout.setContentsMargins(0, 12, 0, 0)
        empty_layout.setSpacing(4)
        empty_icon = QLabel("📭")
        empty_icon.setAlignment(Qt.AlignCenter)
        empty_icon.setStyleSheet("font-size: 32px;")
        empty_label = QLabel("暂无最近操作记录")
        empty_label.setAlignment(Qt.AlignCenter)
        empty_label.setStyleSheet("font-size: 14px; font-weight: 600; color: #9CA3AF;")
        empty_hint = QLabel("你的登录和安全操作将在这里显示")
        empty_hint.setAlignment(Qt.AlignCenter)
        empty_hint.setStyleSheet("font-size: 12px; color: #D1D5DB;")
        empty_layout.addWidget(empty_icon)
        empty_layout.addWidget(empty_label)
        empty_layout.addWidget(empty_hint)
        card_layout.addWidget(self.profile_empty_state)

        self.profile_activity_table.hide()
        self.profile_empty_state.show()

        return card

    def _draw_avatar(self, letter: str):
        size = 72
        pixmap = QPixmap(size, size)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)

        bg_colors = {
            "A": "#3b82f6", "B": "#8b5cf6", "C": "#06b6d4", "D": "#f59e0b",
            "E": "#ef4444", "F": "#10b981", "G": "#6366f1", "H": "#ec4899",
            "J": "#f97316", "K": "#14b8a6", "L": "#84cc16", "M": "#a855f7",
            "N": "#0ea5e9", "P": "#d946ef", "Q": "#64748b", "R": "#f43f5e",
            "S": "#22c55e", "T": "#7c3aed", "W": "#0891b2", "X": "#dc2626",
            "Y": "#ca8a04", "Z": "#9333ea",
        }
        color_name = bg_colors.get(letter.upper(), "#3b82f6")
        bg_color = QColor(color_name)
        painter.setBrush(QBrush(bg_color))
        painter.setPen(Qt.NoPen)
        painter.drawEllipse(4, 4, size - 8, size - 8)

        font = QFont("Microsoft YaHei", 28, QFont.Bold)
        painter.setFont(font)
        painter.setPen(QPen(QColor("#ffffff")))
        painter.drawText(0, 0, size, size, Qt.AlignCenter, letter.upper())

        painter.end()
        self.profile_avatar.setPixmap(pixmap)


    def _build_password_page(self):
        page = EmployeeManagementPage(self.current_user, on_permission_changed=self.refresh_current_user_permissions)
        self.employee_page_widget = page
        return page

    def _switch_page(self, key: str):
        page_keys = list(self.page_widgets.keys())
        if key not in self.page_widgets:
            return
        if key == "password" and not self._has_permission("can_manage_users"):
            QMessageBox.warning(self, "权限不足", "员工管理页面仅管理员或拥有'管理员工'权限的用户可访问。")
            return
        index = page_keys.index(key)
        self.pages.setCurrentIndex(index)
        for name, button in self.sidebar_buttons.items():
            button.setChecked(name == key)
        title, subtitle = self.page_titles.get(key, ("系统首页", ""))
        self.page_title_label.setText(title)
        self.page_desc_label.setText(subtitle)
        self._refresh_header_user_info()
        if key == "statistics":
            self._refresh_statistics_sources()
            self.apply_statistics_filters()
        if key == "profile":
            self._refresh_profile_page()
        if key == "password" and self._has_permission("can_manage_users"):
            if hasattr(self, "employee_page_widget"):
                self.employee_page_widget.current_user = self.current_user
                self.employee_page_widget.refresh_list()
        self.statusBar().showMessage(f"当前页面：{title}")

    def _refresh_header_user_info(self):
        username = (self.current_user or {}).get("username", "未登录")
        role = (self.current_user or {}).get("role", "管理员")
        self.user_badge.set_value(username)
        self.role_badge.set_value(role)

    def _apply_dashboard_style(self):
        self.setStyleSheet(
            """
            QMainWindow, QWidget#MainWindow {
                background: #eef4fb;
            }
            QWidget#Sidebar {
                background: linear-gradient(180deg, #173a63 0%, #1f4c82 100%);
            }
            QLabel#SidebarBrandTitle {
                color: #0f172a;
                font-size: 22px;
                font-weight: 900;
            }
            QLabel#SidebarBrandSubtitle {
                color: #64748b;
                font-size: 14px;
            }
            QPushButton {
                border: none;
                border-radius: 10px;
                padding: 10px 14px;
                background: #ffffff;
                color: #22436a;
            }
            QPushButton:hover {
                background: #e8f1ff;
            }
            QPushButton:pressed {
                background: #d9e7fb;
            }
            QPushButton:checked {
                background: #2f80ed;
                color: #ffffff;
                font-weight: 600;
            }
            QPushButton#SidebarLogoutButton {
                background: rgba(255,255,255,0.15);
                color: #ffffff;
                border: 1px solid rgba(255,255,255,0.22);
            }
            QWidget#PageHeader {
                background: rgba(255, 255, 255, 0.7);
                border: 1px solid rgba(255, 255, 255, 0.85);
                border-radius: 16px;
            }
            QLabel#PageTitle {
                font-size: 22px;
                font-weight: 700;
                color: #16324f;
            }
            QLabel#PageSubtitle {
                color: #6a7a90;
            }
            QFrame[card="true"] {
                background: #ffffff;
                border: 1px solid #dbe6f2;
                border-radius: 16px;
            }
            QLabel#DashboardSectionTitle {
                font-size: 16px;
                font-weight: 700;
                color: #16324f;
            }
            QLabel#DashboardSectionSubtitle {
                color: #70839b;
            }
            QLabel#DashboardMetricTitle {
                color: #6d7f95;
                font-size: 12px;
            }
            QLabel#DashboardMetricValue {
                color: #16324f;
                font-size: 28px;
                font-weight: 700;
                padding: 4px 0px;
            }
            QLabel#DashboardMetricHint {
                color: #8a9ab0;
                font-size: 11px;
            }
            QLabel#UserBadgeTitle {
                color: #6d7f95;
                font-size: 12px;
            }
            QLabel#UserBadgeValue {
                color: #16324f;
                font-size: 14px;
                font-weight: 600;
            }
            QTabWidget::pane {
                border: 1px solid #dbe6f2;
                border-radius: 14px;
                background: white;
            }
            QTabBar::tab {
                background: #edf3fb;
                color: #53667d;
                padding: 10px 16px;
                margin-right: 4px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
            }
            QTabBar::tab:selected {
                background: #ffffff;
                color: #2f80ed;
                font-weight: 600;
            }
            QTextEdit, QPlainTextEdit, QLineEdit, QTableWidget {
                border: 1px solid #dbe6f2;
                border-radius: 12px;
                background: #ffffff;
            }
            """
        )

    def _init_menus(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu("文件")
        run_menu = menubar.addMenu("运行")
        config_menu = menubar.addMenu("配置")
        help_menu = menubar.addMenu("帮助")

        self.action_open_video = QAction("打开视频", self)
        self.action_load_config = QAction("加载配置", self)
        self.action_exit = QAction("退出", self)
        file_menu.addAction(self.action_open_video)
        file_menu.addAction(self.action_load_config)
        file_menu.addSeparator()
        file_menu.addAction(self.action_exit)

        self.action_start = QAction("开始", self)
        self.action_pause = QAction("暂停/继续", self)
        self.action_stop = QAction("停止", self)
        run_menu.addAction(self.action_start)
        run_menu.addAction(self.action_pause)
        run_menu.addAction(self.action_stop)

        self.action_save_yaml = QAction("保存YAML", self)
        self.action_load_yaml = QAction("加载YAML", self)
        config_menu.addAction(self.action_save_yaml)
        config_menu.addAction(self.action_load_yaml)

        self.action_about = QAction("关于", self)
        help_menu.addAction(self.action_about)

        self.action_open_video.triggered.connect(self.control_panel.select_video)
        self.action_load_config.triggered.connect(self.load_yaml_config)
        self.action_exit.triggered.connect(self.close)
        self.action_start.triggered.connect(self.start_processing)
        self.action_pause.triggered.connect(self.toggle_pause)
        self.action_stop.triggered.connect(self.stop_processing)
        self.action_save_yaml.triggered.connect(self.save_yaml_config)
        self.action_load_yaml.triggered.connect(self.load_yaml_config)
        self.action_about.triggered.connect(self.show_about)

    def connect_signals(self):
        self.control_panel.btn_start.clicked.connect(self.start_processing)
        self.control_panel.btn_pause.clicked.connect(self.toggle_pause)
        self.control_panel.btn_stop.clicked.connect(self.stop_processing)
        self.control_panel.btn_select_camera.clicked.connect(self.use_camera_source)
        self.control_panel.btn_export.clicked.connect(self.on_export_results_requested)

        self._reconnect_worker_signals()

        self._stale_timer = QTimer(self)
        self._stale_timer.setInterval(5000)
        self._stale_timer.timeout.connect(self._check_stale_data)

        self.annotation_panel.params_changed.connect(self.apply_runtime_params)
        self.annotation_panel.draw_mode_changed.connect(self._on_draw_mode_changed)
        self.annotation_panel.clear_requested.connect(self.video_panel.clear_annotations)
        self.annotation_panel.undo_requested.connect(self.video_panel.undo_last_annotation_point)
        self.annotation_panel.load_config_requested.connect(self.load_yaml_config)
        self.annotation_panel.save_config_requested.connect(self.save_yaml_config)

        self.video_panel.roi_changed.connect(self.on_annotations_changed)
        self.video_panel.line_changed.connect(self.on_annotations_changed)

        self.video_player.play_pause_clicked.connect(self.toggle_pause)
        self.video_player.seek_requested.connect(self._on_seek_requested)
        self.video_player.step_forward.connect(self._on_step_forward)
        self.video_player.step_backward.connect(self._on_step_backward)
        self._last_frame_position = 0

        self.control_panel.video_selected.connect(self.on_video_selected)
        self.control_panel.save_frame_requested.connect(self.on_save_frame_requested)
        # 模型导出/量化/benchmark 信号
        self.control_panel.export_onnx_requested.connect(self.on_export_onnx_requested)
        self.control_panel.quantize_onnx_requested.connect(self.on_quantize_onnx_requested)
        self.control_panel.benchmark_onnx_requested.connect(self.on_benchmark_onnx_requested)

    def _is_admin(self):
        return (self.current_user or {}).get("role") == "管理员"

    def _has_permission(self, permission_name: str) -> bool:
        perms = (self.current_user or {}).get("permissions", {})
        if not perms:
            return self._is_admin()
        return perms.get(permission_name, False)

    def _check_permission(self, permission_name: str, action_name: str = "") -> bool:
        if self._has_permission(permission_name):
            return True
        display_name = action_name or permission_name
        QMessageBox.warning(self, "权限不足", f"您没有执行此操作的权限（{display_name}），如需开通请联系管理员。")
        return False

    def _require_admin(self, action_name: str) -> bool:
        if self._is_admin():
            return True
        QMessageBox.warning(self, "权限不足", f"员工身份无法执行{action_name}。")
        return False

    def apply_user_permissions(self):
        user_perms = (self.current_user or {}).get("permissions", {})

        can_switch = self._has_permission("can_switch_model")
        can_edit_roi = self._has_permission("can_edit_roi")
        can_edit_line = self._has_permission("can_edit_line")
        can_export = self._has_permission("can_export_data")
        can_manage_users = self._has_permission("can_manage_users")
        can_clear_logs = self._has_permission("can_clear_logs")

        switch_model_widgets = (
            self.control_panel.btn_export_onnx,
            self.control_panel.btn_quantize,
            self.control_panel.btn_benchmark,
            self.btn_export_onnx_page,
            self.btn_quantize_page,
            self.btn_benchmark_page,
        )
        for w in switch_model_widgets:
            w.setEnabled(can_switch)

        if hasattr(self.control_panel, "btn_export_onnx"):
            if not can_switch:
                self.control_panel.btn_export_onnx.setToolTip("无权操作：需要 切换模型 权限")
                self.control_panel.btn_quantize.setToolTip("无权操作：需要 切换模型 权限")
                self.control_panel.btn_benchmark.setToolTip("无权操作：需要 切换模型 权限")

        annotation_widgets = (
            self.annotation_panel.btn_draw_roi,
            self.annotation_panel.btn_draw_line,
            self.annotation_panel.btn_undo,
            self.annotation_panel.btn_clear,
        )
        for w in annotation_widgets:
            w.setEnabled(can_edit_roi or can_edit_line)

        self.annotation_panel.btn_draw_roi.setEnabled(can_edit_roi)
        self.annotation_panel.btn_draw_line.setEnabled(can_edit_line)
        if not can_edit_roi:
            self.annotation_panel.btn_draw_roi.setToolTip("无权操作：需要 编辑ROI 权限")
        if not can_edit_line:
            self.annotation_panel.btn_draw_line.setToolTip("无权操作：需要 编辑计数线 权限")

        export_widgets = (
            self.btn_export_event_table,
            self.btn_export_stats_result,
            self.btn_export_chart_snapshot,
            self.btn_export_history_data,
        )
        for w in export_widgets:
            w.setEnabled(can_export)
        if hasattr(self.control_panel, "btn_export"):
            self.control_panel.btn_export.setEnabled(can_export)
        if not can_export:
            for w in export_widgets:
                w.setToolTip("无权操作：需要 导出数据 权限")

        if hasattr(self, "sidebar_buttons") and self.sidebar_buttons.get("password") is not None:
            self.sidebar_buttons["password"].setEnabled(can_manage_users)

        config_widgets = (
            self.action_load_config,
            self.action_load_yaml,
            self.action_save_yaml,
            self.btn_import_config_page,
            self.btn_save_config_page,
            self.btn_restore_default_config,
            self.btn_view_config_summary,
            self.btn_open_output_dir,
            self.btn_open_log_dir,
            self.btn_clear_temp_files,
            self.btn_view_system_log,
        )
        for w in config_widgets:
            w.setEnabled(can_manage_users or self._is_admin())

        if hasattr(self, "btn_clear_cache"):
            self.btn_clear_cache.setEnabled(can_clear_logs or self._is_admin())

        if hasattr(self, "btn_add_model_page"):
            self.btn_add_model_page.setEnabled(can_switch)
        if hasattr(self, "btn_activate_model_page"):
            self.btn_activate_model_page.setEnabled(can_switch)
        if hasattr(self, "btn_remove_model_page"):
            self.btn_remove_model_page.setEnabled(can_switch)
        if hasattr(self, "btn_evaluate_model_page"):
            self.btn_evaluate_model_page.setEnabled(can_switch)

        role = (self.current_user or {}).get("role", "管理员")
        username = (self.current_user or {}).get("username", "未知用户")
        self.setWindowTitle(f"视频行人客流量统计与分析系统 - {username}（{role}）")
        self.statusBar().showMessage(f"当前用户：{username} | 身份：{role}")
        self._refresh_profile_page()
        self._refresh_management_page()

    def _on_draw_mode_changed(self, mode: str):
        if mode == "roi" and not self._has_permission("can_edit_roi"):
            QMessageBox.warning(self, "权限不足", "您没有绘制/编辑ROI区域的权限，如需开通请联系管理员。")
            return
        if mode == "line" and not self._has_permission("can_edit_line"):
            QMessageBox.warning(self, "权限不足", "您没有绘制/编辑计数线的权限，如需开通请联系管理员。")
            return
        self.video_panel.set_draw_mode(mode)

    def refresh_current_user_permissions(self):
        user_id = (self.current_user or {}).get("id")
        if user_id and self.auth_manager:
            perms = self.auth_manager.get_user_permissions(user_id)
            self.current_user["permissions"] = perms
            self.apply_user_permissions()

    def _refresh_management_page(self):
        config_path = self.current_config_path or "未加载配置"
        detector_cfg = self.current_config.get("detector", {}) if isinstance(self.current_config, dict) else {}
        active_model = get_active_model(self.model_registry) if isinstance(self.model_registry, dict) else None
        model_path = str((active_model or {}).get("path") or detector_cfg.get("model_path", "-"))
        model_name = str((active_model or {}).get("name") or (Path(model_path).stem if model_path and model_path != "-" else "-"))
        model_type = str((active_model or {}).get("framework") or self._detector_type_label())
        model_version = str((active_model or {}).get("version") or detector_cfg.get("version", "-") or "-")
        model_status = str((active_model or {}).get("status") or "未注册")
        model_count = len(self.model_registry.get("models", [])) if isinstance(self.model_registry, dict) else 0
        last_eval = str((active_model or {}).get("last_eval_at") or "-")
        last_metrics = (active_model or {}).get("last_eval_metrics") or {}
        if last_metrics:
            last_metric_text = (
                f"P={float(last_metrics.get('precision', last_metrics.get('metrics/precision(B)', 0.0))):.3f} | "
                f"R={float(last_metrics.get('recall', last_metrics.get('metrics/recall(B)', 0.0))):.3f} | "
                f"mAP50={float(last_metrics.get('map50', last_metrics.get('metrics/mAP50(B)', 0.0))):.3f} | "
                f"mAP50-95={float(last_metrics.get('map5095', last_metrics.get('metrics/mAP50-95(B)', 0.0))):.3f}"
            )
        else:
            last_metric_text = "-"
        input_size = str(detector_cfg.get("imgsz", detector_cfg.get("input_size", 800)))
        current_device = self.current_device or self._device_label()

        if hasattr(self, "management_model_path_label"):
            model_path_display = Path(model_path).name if model_path and model_path != "-" else model_path
            self.management_model_path_label.setText(model_path_display)
            self.management_model_path_label.setToolTip(model_path)
        if hasattr(self, "management_model_name_label"):
            self.management_model_name_label.setText(model_name)
        if hasattr(self, "management_model_type_label"):
            self.management_model_type_label.setText(model_type)
        if hasattr(self, "management_model_version_label"):
            self.management_model_version_label.setText(model_version)
        if hasattr(self, "management_model_status_label"):
            self.management_model_status_label.setText(model_status)
        if hasattr(self, "management_model_count_label"):
            self.management_model_count_label.setText(str(model_count))
        if hasattr(self, "management_model_eval_label"):
            self.management_model_eval_label.setText(f"{last_eval} | {last_metric_text}")
        if hasattr(self, "management_imgsz_label"):
            self.management_imgsz_label.setText(input_size)
        if hasattr(self, "management_device_label"):
            self.management_device_label.setText(current_device)

        if hasattr(self, "management_config_path_label"):
            self.management_config_path_label.setText(config_path)
        if hasattr(self, "management_config_summary_label"):
            self.management_config_summary_label.setText(self._build_config_summary_text())

        if hasattr(self, "management_python_label"):
            self.management_python_label.setText(platform.python_version())
        if hasattr(self, "management_torch_label"):
            self.management_torch_label.setText(self._detect_torch_status())
        if hasattr(self, "management_opencv_label"):
            self.management_opencv_label.setText(self._detect_opencv_status())
        if hasattr(self, "management_cuda_label"):
            self.management_cuda_label.setText("可用" if self._detect_cuda_available() else "不可用")
        if hasattr(self, "management_run_device_label"):
            self.management_run_device_label.setText(current_device)
        if hasattr(self, "management_platform_label"):
            self.management_platform_label.setText(platform.platform())

        self.refresh_model_registry_view()
        self.refresh_model_eval_table()
        self._refresh_home_status_card()

    def _build_config_summary_text(self):
        cfg = self.current_config if isinstance(self.current_config, dict) else self._default_config()
        detector = cfg.get("detector", {})
        alert = cfg.get("alert", self._default_alert_config())
        source = cfg.get("source", "") or "未设置"
        return (
            f"源: {source}\n"
            f"模型: {detector.get('model_path', '-')}\n"
            f"conf: {detector.get('conf', '-')}, iou: {detector.get('iou', '-')}\n"
            f"ROI: {len(cfg.get('counting', {}).get('roi', {}).get('polygon', []))} 点, "
            f"Line: {len(cfg.get('counting', {}).get('line', {}).get('points', []))} 点\n"
            f"告警: {'启用' if alert.get('enabled', False) else '关闭'}, "
            f"桌面通知: {'启用' if alert.get('desktop_notify', True) else '关闭'}, "
            f"总人数阈值 {alert.get('total_threshold', 0)}, "
            f"当前人数阈值 {alert.get('current_threshold', 0)}, "
            f"最低 FPS {float(alert.get('fps_threshold', 0.0)):.1f}"
        )

    def _default_alert_config(self):
        return {
            "enabled": False,
            "desktop_notify": True,
            "total_threshold": 0,
            "current_threshold": 0,
            "fps_threshold": 0.0,
            "cooldown_seconds": 15.0,
        }

    def _normalize_alert_config(self, alert_cfg: dict | None):
        base = self._default_alert_config()
        if isinstance(alert_cfg, dict):
            base.update(alert_cfg)
        try:
            base["enabled"] = bool(base.get("enabled", False))
            base["desktop_notify"] = bool(base.get("desktop_notify", True))
            base["total_threshold"] = max(0, int(base.get("total_threshold", 0) or 0))
            base["current_threshold"] = max(0, int(base.get("current_threshold", 0) or 0))
            base["fps_threshold"] = max(0.0, float(base.get("fps_threshold", 0.0) or 0.0))
            base["cooldown_seconds"] = max(0.0, float(base.get("cooldown_seconds", 15.0) or 0.0))
        except Exception:
            base = self._default_alert_config()
        return base

    def _build_realtime_alert_section(self):
        section = DashboardSection("实时告警", "基于实时统计阈值触发提醒并写入日志")
        body = QWidget()
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(10)

        self.alert_enable_checkbox = QCheckBox("启用实时告警")
        self.alert_desktop_notify_checkbox = QCheckBox("桌面通知")
        self.alert_total_threshold_spin = QSpinBox()
        self.alert_total_threshold_spin.setRange(0, 100000)
        self.alert_total_threshold_spin.setSingleStep(5)
        self.alert_total_threshold_spin.setSuffix(" 人")

        self.alert_current_threshold_spin = QSpinBox()
        self.alert_current_threshold_spin.setRange(0, 100000)
        self.alert_current_threshold_spin.setSingleStep(5)
        self.alert_current_threshold_spin.setSuffix(" 人")

        self.alert_low_fps_spin = QDoubleSpinBox()
        self.alert_low_fps_spin.setRange(0.0, 120.0)
        self.alert_low_fps_spin.setDecimals(1)
        self.alert_low_fps_spin.setSingleStep(0.5)
        self.alert_low_fps_spin.setSuffix(" FPS")

        self.alert_cooldown_spin = QSpinBox()
        self.alert_cooldown_spin.setRange(0, 3600)
        self.alert_cooldown_spin.setSingleStep(5)
        self.alert_cooldown_spin.setSuffix(" 秒")

        self.alert_status_label = QLabel("未启用")
        self.alert_status_label.setWordWrap(True)

        self.alert_history_list = QListWidget()
        self.alert_history_list.setMinimumHeight(160)

        self.btn_clear_alert_history = QPushButton("清空告警记录")
        self.btn_clear_alert_history.clicked.connect(self.clear_alert_history)

        form = QFormLayout()
        form.addRow("实时告警", self.alert_enable_checkbox)
        form.addRow("桌面通知", self.alert_desktop_notify_checkbox)
        form.addRow("总人数阈值", self.alert_total_threshold_spin)
        form.addRow("当前人数阈值", self.alert_current_threshold_spin)
        form.addRow("最低 FPS", self.alert_low_fps_spin)
        form.addRow("告警冷却", self.alert_cooldown_spin)
        form.addRow("当前状态", self.alert_status_label)
        body_layout.addLayout(form)
        body_layout.addWidget(self.alert_history_list)
        body_layout.addWidget(self.btn_clear_alert_history)
        section.set_body_widget(body)

        self.alert_enable_checkbox.stateChanged.connect(self._sync_alert_settings_from_ui)
        self.alert_desktop_notify_checkbox.stateChanged.connect(self._sync_alert_settings_from_ui)
        self.alert_total_threshold_spin.valueChanged.connect(self._sync_alert_settings_from_ui)
        self.alert_current_threshold_spin.valueChanged.connect(self._sync_alert_settings_from_ui)
        self.alert_low_fps_spin.valueChanged.connect(self._sync_alert_settings_from_ui)
        self.alert_cooldown_spin.valueChanged.connect(self._sync_alert_settings_from_ui)

        self._apply_alert_config_to_ui(self.current_config.get("alert", self._default_alert_config()))
        return section

    def _toggle_alert_section(self, checked):
        if hasattr(self, "realtime_alert_section"):
            self.realtime_alert_section.setVisible(checked)
        if hasattr(self, "alert_toggle_btn"):
            self.alert_toggle_btn.setText("▼ 实时告警设置" if checked else "▶ 实时告警设置")

    def _collect_alert_config_from_ui(self):
        if not hasattr(self, "alert_enable_checkbox"):
            return self._default_alert_config()
        return self._normalize_alert_config({
            "enabled": self.alert_enable_checkbox.isChecked(),
            "desktop_notify": self.alert_desktop_notify_checkbox.isChecked(),
            "total_threshold": self.alert_total_threshold_spin.value(),
            "current_threshold": self.alert_current_threshold_spin.value(),
            "fps_threshold": self.alert_low_fps_spin.value(),
            "cooldown_seconds": self.alert_cooldown_spin.value(),
        })

    def _apply_alert_config_to_ui(self, alert_cfg: dict | None):
        alert_cfg = self._normalize_alert_config(alert_cfg)
        if isinstance(self.current_config, dict):
            self.current_config["alert"] = dict(alert_cfg)
        if not hasattr(self, "alert_enable_checkbox"):
            return

        widgets = (
            self.alert_enable_checkbox,
            self.alert_desktop_notify_checkbox,
            self.alert_total_threshold_spin,
            self.alert_current_threshold_spin,
            self.alert_low_fps_spin,
            self.alert_cooldown_spin,
        )
        for widget in widgets:
            widget.blockSignals(True)

        self.alert_enable_checkbox.setChecked(bool(alert_cfg.get("enabled", False)))
        self.alert_desktop_notify_checkbox.setChecked(bool(alert_cfg.get("desktop_notify", True)))
        self.alert_total_threshold_spin.setValue(int(alert_cfg.get("total_threshold", 0)))
        self.alert_current_threshold_spin.setValue(int(alert_cfg.get("current_threshold", 0)))
        self.alert_low_fps_spin.setValue(float(alert_cfg.get("fps_threshold", 0.0)))
        self.alert_cooldown_spin.setValue(int(float(alert_cfg.get("cooldown_seconds", 15.0))))

        for widget in widgets:
            widget.blockSignals(False)

        self._refresh_alert_controls_enabled()
        self._update_alert_status_label()

    def _refresh_alert_controls_enabled(self, *_args):
        enabled = bool(self.alert_enable_checkbox.isChecked()) if hasattr(self, "alert_enable_checkbox") else False
        for widget in (
            getattr(self, "alert_desktop_notify_checkbox", None),
            getattr(self, "alert_total_threshold_spin", None),
            getattr(self, "alert_current_threshold_spin", None),
            getattr(self, "alert_low_fps_spin", None),
            getattr(self, "alert_cooldown_spin", None),
        ):
            if widget is not None:
                widget.setEnabled(enabled)

    def _sync_alert_settings_from_ui(self, *_args):
        if not isinstance(self.current_config, dict):
            self.current_config = self._default_config()
        self.current_config["alert"] = self._collect_alert_config_from_ui()
        self._refresh_alert_controls_enabled()
        self._update_alert_status_label()
        if self.current_detecting and isinstance(self._last_realtime_stats, dict):
            self._evaluate_realtime_alerts(self._last_realtime_stats)
        self._refresh_management_page()

    def _update_alert_status_label(self, active_keys: set[str] | None = None):
        if not hasattr(self, "alert_status_label"):
            return

        if not self.current_detecting:
            self.alert_status_label.setText("未运行")
            self.alert_status_label.setStyleSheet("color: #6b7280; font-weight: bold;")
            return

        alert_cfg = self._collect_alert_config_from_ui() if hasattr(self, "alert_enable_checkbox") else self._default_alert_config()
        if not alert_cfg.get("enabled", False):
            self.alert_status_label.setText("未启用")
            self.alert_status_label.setStyleSheet("color: #6b7280; font-weight: bold;")
            return

        active_keys = active_keys or set()
        desktop_notify_text = "桌面通知开" if alert_cfg.get("desktop_notify", True) else "桌面通知关"
        if active_keys:
            labels = []
            if "total" in active_keys:
                labels.append("总人数")
            if "current" in active_keys:
                labels.append("当前人数")
            if "fps" in active_keys:
                labels.append("低FPS")
            self.alert_status_label.setText(f"告警中：{'、'.join(labels)} | {desktop_notify_text}")
            self.alert_status_label.setStyleSheet("color: #b42318; font-weight: bold;")
        else:
            self.alert_status_label.setText(f"已启用，当前正常 | {desktop_notify_text}")
            self.alert_status_label.setStyleSheet("color: #047857; font-weight: bold;")

    def _init_system_tray(self):
        self.tray_icon = None
        try:
            if not QSystemTrayIcon.isSystemTrayAvailable():
                return
            icon: QIcon = self.style().standardIcon(QStyle.SP_MessageBoxInformation)
            self.setWindowIcon(icon)
            tray_icon = QSystemTrayIcon(icon, self)
            tray_icon.setToolTip("视频行人客流量统计与分析系统")
            tray_icon.show()
            self.tray_icon = tray_icon
        except Exception:
            self.tray_icon = None

    def _show_desktop_notification(self, title: str, message: str, timeout_ms: int = 5000) -> bool:
        tray_icon = getattr(self, "tray_icon", None)
        if tray_icon is None:
            return False
        try:
            if not tray_icon.isVisible():
                tray_icon.show()
            tray_icon.showMessage(title, message, QSystemTrayIcon.Information, timeout_ms)
            return True
        except Exception:
            return False

    def _append_realtime_alert(self, message: str, desktop_notify: bool = False):
        timestamp = datetime.now().strftime("%H:%M:%S")
        run_id = self.current_run_id or "-"
        line = f"[{timestamp}] 批次 {run_id} | {message}"
        self.alert_history_list.insertItem(0, QListWidgetItem(line))
        while self.alert_history_list.count() > 30:
            self.alert_history_list.takeItem(self.alert_history_list.count() - 1)
        self.log_panel.append_log(f"实时告警: {message}")
        self.statusBar().showMessage(f"实时告警：{message}", 5000)
        if desktop_notify:
            notified = self._show_desktop_notification("实时告警", f"批次 {run_id} | {message}")
            if not notified:
                self.log_panel.append_log("实时告警桌面通知不可用，已保留日志与状态栏提示")

    def _evaluate_realtime_alerts(self, stats: dict):
        if not hasattr(self, "alert_enable_checkbox"):
            return

        alert_cfg = self._collect_alert_config_from_ui()
        if not alert_cfg.get("enabled", False):
            self._active_alert_keys = set()
            self._alert_last_emit_at = {}
            self._update_alert_status_label()
            return

        now = time.time()
        cooldown_seconds = float(alert_cfg.get("cooldown_seconds", 15.0) or 0.0)
        active_keys = set()
        message_map = {}
        total_threshold = int(alert_cfg.get("total_threshold", 0) or 0)
        current_threshold = int(alert_cfg.get("current_threshold", 0) or 0)
        fps_threshold = float(alert_cfg.get("fps_threshold", 0.0) or 0.0)
        total_people = int(stats.get("total", 0) or 0)
        current_people = int(stats.get("current", 0) or 0)
        fps_value = float(stats.get("fps", 0.0) or 0.0)

        if total_threshold > 0 and total_people >= total_threshold:
            active_keys.add("total")
            message_map["total"] = f"累计通行人数 {total_people} 已达到阈值 {total_threshold}"
        if current_threshold > 0 and current_people >= current_threshold:
            active_keys.add("current")
            message_map["current"] = f"当前画面人数 {current_people} 已达到阈值 {current_threshold}"
        if fps_threshold > 0.0 and fps_value <= fps_threshold:
            active_keys.add("fps")
            message_map["fps"] = f"实时 FPS {fps_value:.1f} 低于阈值 {fps_threshold:.1f}"

        previous_keys = getattr(self, "_active_alert_keys", set())
        new_keys = active_keys - previous_keys
        cleared_keys = previous_keys - active_keys

        firing_keys = []
        for key in sorted(new_keys):
            last_emit = float(self._alert_last_emit_at.get(key, 0.0) or 0.0)
            if cooldown_seconds <= 0.0 or (now - last_emit) >= cooldown_seconds:
                firing_keys.append(key)
                self._alert_last_emit_at[key] = now

        if firing_keys:
            self._append_realtime_alert(
                "；".join(message_map[key] for key in firing_keys if key in message_map),
                desktop_notify=bool(alert_cfg.get("desktop_notify", True)),
            )
        elif not active_keys and cleared_keys:
            self.log_panel.append_log("实时告警已解除")
            self.statusBar().showMessage("实时告警已解除", 3000)

        self._active_alert_keys = active_keys
        self._update_alert_status_label(active_keys)

    def clear_alert_history(self, *_args, silent: bool = False):
        if hasattr(self, "alert_history_list"):
            self.alert_history_list.clear()
        if not silent:
            self.log_panel.append_log("已清空实时告警记录")

    def _detect_torch_status(self):
        try:
            import torch
            return f"已安装，版本 {torch.__version__}"
        except Exception:
            return "未安装"

    def _detect_opencv_status(self):
        try:
            import cv2
            return f"已安装，版本 {cv2.__version__}"
        except Exception:
            return "未安装"

    def _detect_cuda_available(self):
        try:
            import torch
            return bool(torch.cuda.is_available())
        except Exception:
            return False

    def _load_sample_management_benchmark_records(self):
        self._management_benchmark_records = [
            {"time": "2026-05-06 09:30:00", "model": "best.pt", "imgsz": "800", "fps": "28.4", "latency": "35.2 ms", "device": "CPU", "note": "示例记录"},
            {"time": "2026-05-06 09:42:00", "model": "best.pt", "imgsz": "960", "fps": "24.6", "latency": "40.7 ms", "device": "GPU", "note": "示例记录"},
        ]

    def _load_sample_management_logs(self):
        self._management_logs = [
            {"time": "2026-05-06 09:45:00", "type": "配置", "content": "加载默认配置", "result": "成功", "note": ""},
            {"time": "2026-05-06 09:46:30", "type": "模型", "content": "刷新模型信息", "result": "成功", "note": ""},
        ]

    def refresh_management_benchmark_table(self):
        if not hasattr(self, "management_benchmark_table"):
            return
        rows = getattr(self, "_management_benchmark_records", [])
        self.management_benchmark_table.setRowCount(0)
        for row in rows:
            row_index = self.management_benchmark_table.rowCount()
            self.management_benchmark_table.insertRow(row_index)
            values = [row.get("time", ""), row.get("model", ""), row.get("imgsz", ""), row.get("fps", ""), row.get("latency", ""), row.get("device", ""), row.get("note", "")]
            for col, value in enumerate(values):
                self.management_benchmark_table.setItem(row_index, col, QTableWidgetItem(str(value)))

    def refresh_management_log_table(self):
        if not hasattr(self, "management_log_table"):
            return
        rows = getattr(self, "_management_logs", [])
        self.management_log_table.setRowCount(0)
        for row in rows:
            row_index = self.management_log_table.rowCount()
            self.management_log_table.insertRow(row_index)
            values = [row.get("time", ""), row.get("type", ""), row.get("content", ""), row.get("result", ""), row.get("note", "")]
            for col, value in enumerate(values):
                self.management_log_table.setItem(row_index, col, QTableWidgetItem(str(value)))

    def _append_management_log(self, log_type: str, content: str, result: str = "成功", note: str = ""):
        log_item = {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type": log_type,
            "content": content,
            "result": result,
            "note": note,
        }
        self._management_logs = [log_item] + getattr(self, "_management_logs", [])[:29]
        self.refresh_management_log_table()

    def _append_management_benchmark_record(self, model_name: str, input_size: str, device: str, note: str = ""):
        record = {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "model": model_name,
            "imgsz": input_size,
            "fps": "-",
            "latency": "-",
            "device": device,
            "note": note,
        }
        self._management_benchmark_records = [record] + getattr(self, "_management_benchmark_records", [])[:19]
        self.refresh_management_benchmark_table()

    def _default_eval_device_arg(self):
        try:
            import torch

            if torch.cuda.is_available():
                return 0
        except Exception:
            pass
        return "cpu"

    def _format_metric_value(self, value, precision: int = 3):
        try:
            return f"{float(value):.{precision}f}"
        except Exception:
            return "-"

    def _format_model_config_summary(self, record: dict) -> str:
        conf = self._format_metric_value(record.get("conf"))
        iou = self._format_metric_value(record.get("iou"))
        imgsz = record.get("imgsz", "-") or "-"
        device = record.get("device", "-") or "-"
        return f"conf={conf} | iou={iou} | imgsz={imgsz} | device={device}"

    def _format_model_metrics_summary(self, metrics: dict) -> str:
        if not isinstance(metrics, dict) or not metrics:
            return "-"
        precision = metrics.get("precision", metrics.get("metrics/precision(B)", 0.0))
        recall = metrics.get("recall", metrics.get("metrics/recall(B)", 0.0))
        map50 = metrics.get("map50", metrics.get("metrics/mAP50(B)", 0.0))
        map5095 = metrics.get("map5095", metrics.get("metrics/mAP50-95(B)", 0.0))
        return (
            f"P={self._format_metric_value(precision)} | "
            f"R={self._format_metric_value(recall)} | "
            f"mAP50={self._format_metric_value(map50)} | "
            f"mAP50-95={self._format_metric_value(map5095)}"
        )

    def _selected_model_record(self) -> dict | None:
        table = getattr(self, "management_model_registry_table", None)
        if table is None:
            return None
        selection_model = table.selectionModel()
        if selection_model is None:
            return None
        selected_rows = selection_model.selectedRows()
        if not selected_rows:
            return None
        row_index = selected_rows[0].row()
        item = table.item(row_index, 0)
        model_id = item.data(Qt.UserRole) if item is not None else ""
        if not model_id:
            return None
        return get_model_by_id(self.model_registry, model_id)

    def refresh_model_registry_view(self):
        table = getattr(self, "management_model_registry_table", None)
        if table is None:
            return
        records = list(self.model_registry.get("models", [])) if isinstance(self.model_registry, dict) else []
        table.setRowCount(0)
        for record in records:
            row_index = table.rowCount()
            table.insertRow(row_index)
            status = "激活" if str(record.get("status", "")) == "active" else "未激活"
            values = [
                status,
                str(record.get("name", "")),
                str(record.get("version", "")),
                str(record.get("framework", "")),
                str(record.get("path", "")),
                self._format_model_config_summary(record),
                str(record.get("evaluation_count", 0)),
                str(record.get("last_eval_at", "-")) or "-",
                self._format_model_metrics_summary(record.get("last_eval_metrics", {})),
                str(record.get("updated_at", "-")) or "-",
                str(record.get("notes", "")),
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if col == 0:
                    item.setData(Qt.UserRole, str(record.get("id", "")))
                table.setItem(row_index, col, item)

    def refresh_model_eval_table(self):
        table = getattr(self, "management_model_eval_table", None)
        if table is None:
            return
        records = list(self.model_registry.get("evaluations", [])) if isinstance(self.model_registry, dict) else []
        table.setRowCount(0)
        for record in records:
            row_index = table.rowCount()
            table.insertRow(row_index)
            metrics = dict(record.get("metrics", {}) or {})
            values = [
                str(record.get("created_at", "")),
                str(record.get("type", "")),
                str(record.get("model_name", "")),
                str(record.get("version", "")),
                str(record.get("dataset", "")),
                str(record.get("device", "-")),
                self._format_metric_value(metrics.get("precision", metrics.get("metrics/precision(B)", 0.0))),
                self._format_metric_value(metrics.get("recall", metrics.get("metrics/recall(B)", 0.0))),
                self._format_metric_value(metrics.get("map50", metrics.get("metrics/mAP50(B)", 0.0))),
                self._format_metric_value(metrics.get("map5095", metrics.get("metrics/mAP50-95(B)", 0.0))),
                self._format_metric_value(metrics.get("fps", record.get("fps", 0.0))),
                f"{self._format_metric_value(record.get('duration_seconds', metrics.get('elapsed_seconds', 0.0)))} s",
                str(record.get("note", "")),
            ]
            for col, value in enumerate(values):
                table.setItem(row_index, col, QTableWidgetItem(str(value)))

    def _sync_model_registry_from_current_config(self, save: bool = True):
        detector_cfg = self.current_config.get("detector", {}) if isinstance(self.current_config, dict) else {}
        model_path = str(detector_cfg.get("model_path", "")).strip()
        if not model_path:
            return None

        existing = get_model_by_path(self.model_registry, model_path)
        if existing is None:
            record = ensure_model_record(
                self.model_registry,
                model_path,
                name=Path(model_path).stem,
                version=str(detector_cfg.get("version", "current") or "current"),
                notes="当前配置模型",
                conf=float(detector_cfg.get("conf", 0.5)),
                iou=float(detector_cfg.get("iou", 0.45)),
                imgsz=int(detector_cfg.get("imgsz", detector_cfg.get("input_size", 800)) or 800),
                device=detector_cfg.get("device", self._default_eval_device_arg()),
                framework=infer_framework(model_path),
                make_active=True,
            )
        else:
            if detector_cfg.get("conf") is not None:
                existing["conf"] = float(detector_cfg.get("conf", existing.get("conf", 0.5)))
            if detector_cfg.get("iou") is not None:
                existing["iou"] = float(detector_cfg.get("iou", existing.get("iou", 0.45)))
            if detector_cfg.get("imgsz") is not None or detector_cfg.get("input_size") is not None:
                existing["imgsz"] = int(detector_cfg.get("imgsz", detector_cfg.get("input_size", existing.get("imgsz", 800))) or 800)
            if detector_cfg.get("device") is not None:
                existing["device"] = detector_cfg.get("device")
            existing["framework"] = existing.get("framework") or infer_framework(model_path)
            existing["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            record = existing
        activate_model(self.model_registry, record["id"])
        if save:
            save_registry(self.model_registry)
        self.refresh_model_registry_view()
        self.refresh_model_eval_table()
        return record

    def add_model_version(self):
        if not self._check_permission("can_switch_model", "添加模型版本"):
            return
        current_model_path = str(self.current_config.get("detector", {}).get("model_path", "")) if isinstance(self.current_config, dict) else ""
        default_dir = str(Path(current_model_path).parent) if current_model_path else str(writable_path("runs/train"))
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择模型文件",
            default_dir,
            "Model Files (*.pt *.pth *.onnx *.engine *.torchscript *.ts);;All Files (*)",
        )
        if not path:
            return

        default_name = Path(path).stem
        name, ok = QInputDialog.getText(self, "添加模型版本", "模型名称", QLineEdit.Normal, default_name)
        if not ok:
            return
        version_default = f"v{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        version, ok = QInputDialog.getText(self, "添加模型版本", "版本号", QLineEdit.Normal, version_default)
        if not ok:
            return
        notes, _ = QInputDialog.getText(self, "添加模型版本", "备注（可选）", QLineEdit.Normal, "")

        record = ensure_model_record(
            self.model_registry,
            path,
            name=name or default_name,
            version=version or version_default,
            notes=notes or "",
            conf=float(self.current_config.get("detector", {}).get("conf", 0.5)) if isinstance(self.current_config, dict) else 0.5,
            iou=float(self.current_config.get("detector", {}).get("iou", 0.45)) if isinstance(self.current_config, dict) else 0.45,
            imgsz=int(self.current_config.get("detector", {}).get("imgsz", self.current_config.get("detector", {}).get("input_size", 800)) or 800) if isinstance(self.current_config, dict) else 800,
            device=self._default_eval_device_arg(),
            framework=infer_framework(path),
            make_active=False,
        )
        save_registry(self.model_registry)
        self.refresh_model_registry_view()
        self._refresh_management_page()
        self._append_management_log("模型", f"添加模型版本: {record['name']}@{record['version']}", "成功", notes or "")

    def activate_selected_model(self):
        if not self._check_permission("can_switch_model", "激活模型"):
            return
        record = self._selected_model_record()
        if record is None:
            QMessageBox.warning(self, "未选择模型", "请先在模型版本表中选中一个模型。")
            return
        if self.worker.isRunning():
            QMessageBox.warning(self, "无法切换", "当前正在检测，请先停止检测后再切换模型。")
            return

        activate_model(self.model_registry, record["id"])
        detector_cfg = self.current_config.setdefault("detector", {})
        detector_cfg["model_path"] = record.get("path", detector_cfg.get("model_path", ""))
        if record.get("conf") is not None:
            detector_cfg["conf"] = float(record["conf"])
        if record.get("iou") is not None:
            detector_cfg["iou"] = float(record["iou"])
        if record.get("imgsz") is not None:
            detector_cfg["imgsz"] = int(record["imgsz"])
        if record.get("device") is not None:
            detector_cfg["device"] = record["device"]

        self.worker.set_config(self.current_config)
        save_registry(self.model_registry)
        self._refresh_system_info()
        self._refresh_management_page()
        self._append_management_log("模型", f"激活模型: {record.get('name', '模型')}@{record.get('version', '')}", "成功")

    def remove_selected_model(self):
        if not self._check_permission("can_switch_model", "删除模型"):
            return
        record = self._selected_model_record()
        if record is None:
            QMessageBox.warning(self, "未选择模型", "请先在模型版本表中选中一个模型。")
            return
        if len(self.model_registry.get("models", [])) <= 1:
            QMessageBox.warning(self, "无法删除", "至少需要保留一个模型版本。")
            return
        result = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除模型 {record.get('name', '模型')}@{record.get('version', '')} 吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if result != QMessageBox.Yes:
            return
        if self.worker.isRunning() and str(record.get("status", "")) == "active":
            QMessageBox.warning(self, "无法删除", "当前正在检测，不能删除正在使用的模型。")
            return

        removed = remove_model(self.model_registry, record["id"])
        save_registry(self.model_registry)
        active_model = get_active_model(self.model_registry)
        if active_model is not None:
            detector_cfg = self.current_config.setdefault("detector", {})
            detector_cfg["model_path"] = active_model.get("path", detector_cfg.get("model_path", ""))
            if active_model.get("conf") is not None:
                detector_cfg["conf"] = float(active_model["conf"])
            if active_model.get("iou") is not None:
                detector_cfg["iou"] = float(active_model["iou"])
            if active_model.get("imgsz") is not None:
                detector_cfg["imgsz"] = int(active_model["imgsz"])
            if active_model.get("device") is not None:
                detector_cfg["device"] = active_model["device"]
            self.worker.set_config(self.current_config)
        self._refresh_system_info()
        self._refresh_management_page()
        self._append_management_log("模型", f"删除模型: {removed.get('name', '模型')}@{removed.get('version', '')}", "成功")

    def evaluate_selected_model(self):
        if not self._check_permission("can_switch_model", "模型评估"):
            return
        if self.worker.isRunning():
            QMessageBox.warning(self, "无法评估", "当前正在检测，请先停止检测后再执行模型评估。")
            return
        if self._model_eval_thread is not None and self._model_eval_thread.isRunning():
            QMessageBox.information(self, "评估中", "当前已有模型评估任务正在运行。")
            return

        record = self._selected_model_record()
        if record is None:
            QMessageBox.warning(self, "未选择模型", "请先在模型版本表中选中一个模型。")
            return

        default_dataset = str(resource_path("datasets/pedestrian_all/data.yaml"))
        dataset_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择评估数据集",
            str(Path(default_dataset).parent),
            "YAML Files (*.yaml *.yml);;All Files (*)",
        )
        if not dataset_path:
            dataset_path = default_dataset
        if not Path(dataset_path).exists():
            QMessageBox.warning(self, "数据集不存在", f"找不到评估数据集: {dataset_path}")
            return

        detector_cfg = self.current_config.get("detector", {}) if isinstance(self.current_config, dict) else {}
        imgsz = int(record.get("imgsz") or detector_cfg.get("imgsz", detector_cfg.get("input_size", 800)) or 800)
        batch = int(detector_cfg.get("batch", 16) or 16)
        device = record.get("device") or detector_cfg.get("device") or self._default_eval_device_arg()

        self._model_eval_thread = ModelEvaluationThread(
            model_path=str(record.get("path", "")),
            data_path=dataset_path,
            imgsz=imgsz,
            batch=batch,
            device=device,
            split="val",
            augment=False,
            parent=self,
        )
        self._model_eval_thread.log.connect(self.log_panel.append_log)
        self._model_eval_thread.done.connect(lambda payload, ref=record, dataset=dataset_path: self._on_model_evaluation_finished(ref, payload, dataset))
        self._model_eval_thread.failed.connect(self._on_model_evaluation_failed)
        self._model_eval_thread.finished.connect(self._on_model_evaluation_thread_finished)
        self._append_management_log("模型评估", f"开始评估: {record.get('name', '模型')}@{record.get('version', '')}", "进行中", Path(dataset_path).name)
        self.log_panel.append_log(f"开始模型评估: {record.get('name', '模型')}@{record.get('version', '')}")
        self._model_eval_thread.start()

    def _on_model_evaluation_finished(self, record: dict, payload: dict, dataset_path: str):
        metrics = {
            "precision": float(payload.get("precision", 0.0) or 0.0),
            "recall": float(payload.get("recall", 0.0) or 0.0),
            "map50": float(payload.get("map50", 0.0) or 0.0),
            "map5095": float(payload.get("map5095", 0.0) or 0.0),
            "fitness": float(payload.get("fitness", 0.0) or 0.0),
            "fps": float(payload.get("fps", 0.0) or 0.0),
            "inference_ms": float(payload.get("inference_ms", 0.0) or 0.0),
            "elapsed_seconds": float(payload.get("elapsed_seconds", 0.0) or 0.0),
            "imgsz": int(payload.get("imgsz", 0) or 0),
            "batch": int(payload.get("batch", 0) or 0),
            "device": payload.get("device", ""),
        }
        note = f"split={payload.get('split', 'val')}"
        try:
            evaluation_record = update_model_evaluation(
                self.model_registry,
                str(record.get("id", "")),
                "验证",
                dataset_path,
                metrics,
                note=note,
                duration_seconds=float(payload.get("elapsed_seconds", 0.0) or 0.0),
            )
        except KeyError:
            self._append_management_log("模型评估", "评估完成但模型记录已不存在", "失败", str(record.get("name", "")))
            self.log_panel.append_log("模型评估完成，但模型记录已不存在")
            return
        evaluation_record["device"] = str(payload.get("device", ""))
        evaluation_record["imgsz"] = int(payload.get("imgsz", 0) or 0)
        evaluation_record["batch"] = int(payload.get("batch", 0) or 0)
        evaluation_record["fps"] = float(payload.get("fps", 0.0) or 0.0)
        evaluation_record["inference_ms"] = float(payload.get("inference_ms", 0.0) or 0.0)
        save_registry(self.model_registry)
        self.refresh_model_registry_view()
        self.refresh_model_eval_table()
        self._refresh_management_page()
        self._append_management_log(
            "模型评估",
            f"完成评估: {record.get('name', '模型')}@{record.get('version', '')}",
            "成功",
            Path(dataset_path).name,
        )

    def _on_model_evaluation_failed(self, message: str):
        self.log_panel.append_log(f"模型评估失败: {message}")
        self._append_management_log("模型评估", "模型评估任务失败", "失败", message)
        QMessageBox.critical(self, "评估失败", f"模型评估失败: {message}")

    def _on_model_evaluation_thread_finished(self):
        if self._model_eval_thread is not None:
            self._model_eval_thread.deleteLater()
            self._model_eval_thread = None

    def restore_default_config(self):
        if not self._require_admin("恢复默认配置"):
            return
        self.current_config = self._default_config()
        self.current_config_path = None
        self.worker.set_config(self.current_config)
        self._apply_alert_config_to_ui(self.current_config.get("alert", self._default_alert_config()))
        self.annotation_panel.set_params({
            "conf": float(self.current_config.get("detector", {}).get("conf", 0.5)),
            "iou": float(self.current_config.get("detector", {}).get("iou", 0.45)),
            "show_trail": True,
            "show_roi": True,
            "show_line": True,
            "show_heatmap": True,
            "face_blur_enabled": False,
            "draw_count_points": True,
        })
        self._sync_model_registry_from_current_config(save=True)
        self._refresh_system_info()
        self._refresh_management_page()
        self._append_management_log("配置", "恢复默认配置", "成功")

    def show_config_summary(self):
        if not self._require_admin("查看配置摘要"):
            return
        QMessageBox.information(self, "配置摘要", self._build_config_summary_text())

    def open_output_dir(self):
        if not self._require_admin("打开输出目录"):
            return
        output_dir = writable_path("outputs")
        output_dir.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(str(output_dir))
        except Exception:
            QMessageBox.information(self, "打开输出目录", f"输出目录: {output_dir}")
        self._append_management_log("维护", "打开输出目录", "成功")

    def open_log_dir(self):
        if not self._require_admin("打开日志目录"):
            return
        log_dir = writable_path("outputs/gui_run")
        log_dir.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(str(log_dir))
        except Exception:
            QMessageBox.information(self, "打开日志目录", f"日志目录: {log_dir}")
        self._append_management_log("维护", "打开日志目录", "成功")

    def clear_cache(self):
        if not self._require_admin("清理缓存"):
            return
        cleared = 0
        for candidate in [Path(__file__).resolve().parents[1] / "__pycache__", Path(__file__).resolve().parents[1] / "gui" / "__pycache__"]:
            if candidate.exists():
                try:
                    shutil.rmtree(candidate)
                    cleared += 1
                except Exception:
                    pass
        QMessageBox.information(self, "清理缓存", f"已清理缓存目录: {cleared} 个")
        self._append_management_log("维护", "清理缓存", "成功", f"清理 {cleared} 个目录")

    def clear_temp_files(self):
        if not self._require_admin("清空临时文件"):
            return
        temp_dir = writable_path("outputs/gui_run")
        removed = 0
        if temp_dir.exists():
            for path in temp_dir.glob("*.tmp"):
                try:
                    path.unlink()
                    removed += 1
                except Exception:
                    pass
        QMessageBox.information(self, "清空临时文件", f"已清理临时文件: {removed} 个")
        self._append_management_log("维护", "清空临时文件", "成功", f"清理 {removed} 个文件")

    def show_system_log(self):
        if not self._require_admin("查看系统日志"):
            return
        lines = [f"{row.get('time')} | {row.get('type')} | {row.get('content')} | {row.get('result')}" for row in getattr(self, "_management_logs", [])[:10]]
        QMessageBox.information(self, "系统日志", "\n".join(lines) if lines else "暂无系统日志")

    def _refresh_home_status_card(self):
        self.home_status_detect.setText("检测中" if self.current_detecting else "未运行")
        self.home_status_source.setText(self.current_source_name)
        self.home_status_model.setText(self.current_detector_type)
        self.home_status_config.setText(self.current_config_path or "默认配置")

    def _refresh_statistics_sources(self):
        current_text = self.query_video_name.currentText() if hasattr(self, "query_video_name") else "全部视频"
        self.query_video_name.blockSignals(True)
        self.query_video_name.clear()
        self.query_video_name.addItem("全部视频")
        for source_name in self.history_db.query_sources():
            self.query_video_name.addItem(source_name)
        if current_text and self.query_video_name.findText(current_text) >= 0:
            self.query_video_name.setCurrentText(current_text)
        self.query_video_name.blockSignals(False)
        self._refresh_statistics_run_choices()

    def _get_statistics_base_filters(self):
        start_time = self._parse_time_filter(self.query_time_start.text()) if hasattr(self, "query_time_start") else None
        end_time = self._parse_time_filter(self.query_time_end.text()) if hasattr(self, "query_time_end") else None
        if start_time and end_time and end_time < start_time:
            start_time, end_time = end_time, start_time

        selected_video = self.query_video_name.currentText() if hasattr(self, "query_video_name") else "全部视频"
        selected_dir = self.query_direction.currentText() if hasattr(self, "query_direction") else "All"
        bucket_minutes = 1 if hasattr(self, "query_bucket") and self.query_bucket.currentText() == "按分钟" else 5
        mode = self.query_stats_mode.currentText() if hasattr(self, "query_stats_mode") else "最新一次检测"
        return start_time, end_time, selected_video, selected_dir, bucket_minutes, mode

    def _refresh_statistics_run_choices(self, *_args):
        if not hasattr(self, "query_run_batch") or not hasattr(self, "history_db"):
            return

        start_time, end_time, selected_video, _, _, mode = self._get_statistics_base_filters()
        current_data = self.query_run_batch.currentData()
        current_session_id = None
        current_run_id = None
        if isinstance(current_data, dict):
            current_session_id = current_data.get("session_id")
            current_run_id = current_data.get("run_id")

        try:
            run_rows = self.history_db.query_run_batches(start_time=start_time, end_time=end_time, source_name=selected_video)
        except Exception:
            run_rows = []

        self.query_run_batch.blockSignals(True)
        self.query_run_batch.clear()
        if not run_rows:
            self.query_run_batch.addItem("暂无检测批次", {"session_id": None, "run_id": None, "display_text": "暂无检测批次"})
        else:
            for row in run_rows:
                session_id = int(row.get("session_id") or 0)
                run_id = str(row.get("run_id") or session_id or "")
                label = str(row.get("display_text") or f"{row.get('started_at', '')} - {row.get('source_name', '')}")
                self.query_run_batch.addItem(
                    label,
                    {
                        "session_id": session_id,
                        "run_id": run_id,
                        "source_name": row.get("source_name", ""),
                        "video_name": row.get("video_name", ""),
                        "started_at": row.get("started_at", ""),
                        "ended_at": row.get("ended_at", ""),
                        "display_text": label,
                    },
                )

        target_index = 0
        if current_session_id is not None or current_run_id is not None:
            for index in range(self.query_run_batch.count()):
                item_data = self.query_run_batch.itemData(index)
                if not isinstance(item_data, dict):
                    continue
                item_session_id = item_data.get("session_id")
                item_run_id = str(item_data.get("run_id") or item_session_id or "")
                if (current_session_id is not None and str(current_session_id) == str(item_session_id)) or (
                    current_run_id is not None and str(current_run_id) == item_run_id
                ):
                    target_index = index
                    break
        self.query_run_batch.setCurrentIndex(target_index)
        self.query_run_batch.blockSignals(False)
        self.query_run_batch.setEnabled(mode == "指定检测批次")

    def _parse_time_filter(self, text: str):
        raw = (text or "").strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        return None

    def _get_statistics_query_context(self):
        start_time, end_time, selected_video, selected_dir, bucket_minutes, mode = self._get_statistics_base_filters()

        sessions = []
        events = []
        history_rows = []
        fps_rows = []
        selected_run = None
        selected_run_data = {"display_text": "全部历史批次"}
        session_ids = []

        if mode == "历史累计统计":
            sessions = self.history_db.query_sessions(start_time=start_time, end_time=end_time, source_name=selected_video, order="asc")
            session_ids = [int(row.get("id") or 0) for row in sessions if row.get("id")]
            events = self.history_db.query_events(
                start_time=start_time,
                end_time=end_time,
                source_name=selected_video,
                direction=selected_dir,
                session_ids=session_ids,
            )
            history_rows = self.history_db.query_traffic_history(
                start_time=start_time,
                end_time=end_time,
                source_name=selected_video,
                bucket_minutes=bucket_minutes,
                session_ids=session_ids,
            )
            fps_rows = self.history_db.query_fps_history(
                start_time=start_time,
                end_time=end_time,
                source_name=selected_video,
                session_ids=session_ids,
            )
            selected_run_data = {"display_text": f"{len(session_ids)} 个历史批次"}
        else:
            selected_run_data = self._get_selected_statistics_run() if mode == "指定检测批次" else None
            if selected_run_data:
                selected_session_id = int(selected_run_data.get("session_id") or 0)
                if selected_session_id:
                    selected_run_rows = self.history_db.query_sessions(run_id=str(selected_session_id), order="desc", limit=1)
                    selected_run = selected_run_rows[0] if selected_run_rows else None

            if selected_run is None:
                latest_rows = self.history_db.query_sessions(
                    start_time=start_time,
                    end_time=end_time,
                    source_name=selected_video,
                    order="desc",
                    limit=1,
                )
                selected_run = latest_rows[0] if latest_rows else None

            if selected_run:
                session_ids = [int(selected_run.get("id") or 0)]
                sessions = [selected_run]
                selected_run_data = {
                    "session_id": selected_run.get("id", 0),
                    "run_id": selected_run.get("run_id", ""),
                    "display_text": f"{selected_run.get('started_at', '')} - {selected_run.get('source_name', '')}",
                }
            else:
                selected_run_data = {"display_text": "暂无检测批次"}

            events = self.history_db.query_events(
                start_time=start_time,
                end_time=end_time,
                source_name=selected_video,
                direction=selected_dir,
                session_ids=session_ids,
            )
            history_rows = self.history_db.query_traffic_history(
                start_time=start_time,
                end_time=end_time,
                source_name=selected_video,
                bucket_minutes=bucket_minutes,
                session_ids=session_ids,
            )
            fps_rows = self.history_db.query_fps_history(
                start_time=start_time,
                end_time=end_time,
                source_name=selected_video,
                session_ids=session_ids,
            )

        return {
            "mode": mode,
            "start_time": start_time,
            "end_time": end_time,
            "selected_video": selected_video,
            "selected_dir": selected_dir,
            "bucket_minutes": bucket_minutes,
            "events": events,
            "sessions": sessions,
            "history_rows": history_rows,
            "fps_rows": fps_rows,
            "selected_run": selected_run,
            "selected_run_label": (selected_run_data or {}).get("display_text", "暂无检测批次"),
            "session_ids": session_ids,
        }

    def _parse_mmss_to_seconds(self, text: str) -> int:
        raw = (text or "").strip()
        if not raw:
            return 0
        parts = raw.split(":")
        if len(parts) != 2:
            return 0
        try:
            mm = max(0, int(parts[0]))
            ss = max(0, int(parts[1]))
            return mm * 60 + ss
        except ValueError:
            return 0

    def _event_direction(self, event_name: str) -> str:
        low = event_name.lower()
        if "up" in low or "上" in event_name:
            return "Up"
        if "down" in low or "下" in event_name:
            return "Down"
        return "All"

    def _get_selected_statistics_run(self):
        if not hasattr(self, "query_run_batch"):
            return None
        current_data = self.query_run_batch.currentData()
        if isinstance(current_data, dict):
            return current_data
        return None

    def _apply_statistics_mode_labels(self, context):
        mode = context.get("mode", "最新一次检测")
        selected_video = context.get("selected_video", "全部视频")
        selected_dir = context.get("selected_dir", "All")
        selected_run_label = context.get("selected_run_label") or "暂无检测批次"
        session_count = len(context.get("sessions", []))
        event_count = len(context.get("events", []))

        if mode == "历史累计统计":
            total_title = "历史累计总数"
            up_title = "历史累计 Up"
            down_title = "历史累计 Down"
            total_hint = f"{selected_video} · {session_count} 个批次"
            up_hint = f"方向 {selected_dir} · 事件 {event_count} 条"
            down_hint = f"方向 {selected_dir} · 事件 {event_count} 条"
            fps_hint = f"累计样本 · {selected_run_label}"
        else:
            total_title = "本次检测总数"
            up_title = "本次 Up"
            down_title = "本次 Down"
            total_hint = f"{selected_run_label}"
            up_hint = f"方向 {selected_dir} · 当前批次"
            down_hint = f"方向 {selected_dir} · 当前批次"
            fps_hint = f"当前批次 · {selected_run_label}"

        self.stats_query_total.set_title(total_title)
        self.stats_query_up.set_title(up_title)
        self.stats_query_down.set_title(down_title)
        self.stats_query_fps.set_title("平均 FPS")
        self.stats_query_total.set_hint(total_hint)
        self.stats_query_up.set_hint(up_hint)
        self.stats_query_down.set_hint(down_hint)
        self.stats_query_fps.set_hint(fps_hint)

        if hasattr(self, "stats_query_section"):
            self.stats_query_section.set_subtitle(f"模式：{mode} | 视频：{selected_video} | 批次：{selected_run_label}")
        if hasattr(self, "stats_history_section"):
            self.stats_history_section.set_subtitle(f"{mode} · {selected_run_label}")
        if hasattr(self, "stats_fps_section"):
            self.stats_fps_section.set_subtitle(f"{mode} · {selected_run_label}")
        if hasattr(self, "stats_event_section"):
            self.stats_event_section.set_subtitle(f"{mode} · {selected_run_label}")
        if hasattr(self, "stats_experiment_section"):
            self.stats_experiment_section.set_subtitle(f"{mode} · {session_count} 个批次")

    def apply_statistics_filters(self):
        self._refresh_statistics_run_choices()
        context = self._get_statistics_query_context()
        events = context["events"]
        selected_run_label = context.get("selected_run_label", "暂无检测批次")
        mode = context.get("mode", "最新一次检测")

        self.filtered_event_table_panel.set_records([
            {
                "timestamp": item.get("timestamp", ""),
                "event": item.get("event_type", ""),
                "track_id": item.get("track_id", ""),
                "target": item.get("target", ""),
                "run_id": item.get("run_id", ""),
            }
            for item in events
        ])

        up_count = sum(1 for item in events if str(item.get("direction", "")).lower() == "up")
        down_count = sum(1 for item in events if str(item.get("direction", "")).lower() == "down")
        session_rows = context["sessions"]
        total_count = sum(int(row.get("total_count") or 0) for row in session_rows)
        avg_fps = 0.0
        fps_rows = context.get("fps_rows", [])
        fps_values = []
        for row in fps_rows:
            value = row.get("fps") if row.get("fps") is not None else row.get("avg_fps")
            if value is not None:
                fps_values.append(float(value))
        if fps_values:
            avg_fps = sum(fps_values) / len(fps_values)

        self._apply_statistics_mode_labels(context)

        if mode == "历史累计统计":
            self.stats_query_total.set_value(str(total_count), f"{len(session_rows)} 个批次")
            self.stats_query_up.set_value(str(up_count), f"{len(events)} 条事件")
            self.stats_query_down.set_value(str(down_count), f"{len(events)} 条事件")
            self.stats_query_fps.set_value(f"{avg_fps:.2f}", f"{len(fps_values)} 个采样点")
        else:
            self.stats_query_total.set_value(str(total_count), selected_run_label)
            self.stats_query_up.set_value(str(up_count), f"Up · {selected_run_label}")
            self.stats_query_down.set_value(str(down_count), f"Down · {selected_run_label}")
            self.stats_query_fps.set_value(f"{avg_fps:.2f}", f"{selected_run_label}")

        history_points = context["history_rows"]
        self.stats_history_panel.load_history([
            {
                "minute": idx,
                "up": item.get("up", 0),
                "down": item.get("down", 0),
                "total": item.get("total", 0),
            }
            for idx, item in enumerate(history_points)
        ])
        self.stats_fps_panel.load_history([float(row.get("fps") or row.get("avg_fps") or 0.0) for row in fps_rows])

        self.experiment_table_panel.reset()
        for row in session_rows:
            self.experiment_table_panel.add_record(
                timestamp=str(row.get("started_at", "")),
                conf=float(row.get("conf") or 0.0),
                iou=float(row.get("iou") or 0.0),
                avg_fps=float(row.get("avg_fps") or 0.0),
                up=int(row.get("up_count") or 0),
                down=int(row.get("down_count") or 0),
                run_id=str(row.get("run_id") or row.get("id") or ""),
            )

    def reset_statistics_filters(self):
        self.query_time_start.clear()
        self.query_time_end.clear()
        self.query_stats_mode.setCurrentText("最新一次检测")
        self.query_direction.setCurrentText("All")
        self.query_bucket.setCurrentText("按分钟")
        self._refresh_statistics_sources()
        self.query_video_name.setCurrentIndex(0)
        self._refresh_statistics_run_choices()
        self.apply_statistics_filters()

    def _refresh_profile_page(self):
        username = (self.current_user or {}).get("username", "-")
        profile = self.auth_manager.get_user_profile(username) or {}

        def _safe_text(attr, value):
            if hasattr(self, attr):
                try:
                    getattr(self, attr).setText(str(value))
                except Exception:
                    pass

        def _safe_badge(attr, text, bg, fg):
            if hasattr(self, attr):
                try:
                    badge = getattr(self, attr)
                    badge.setText(text)
                    badge.setStyleSheet(
                        "background: {bg}; color: {fg}; font-size: 11px; font-weight: 600;"
                        "padding: 3px 10px; border-radius: 8px;".format(bg=bg, fg=fg)
                    )
                except Exception:
                    pass

        display_name = profile.get("username", username)
        display_role = profile.get("role", (self.current_user or {}).get("role", "管理员"))
        display_status = "已启用" if profile.get("enabled", True) else "已禁用"
        display_login = str(profile.get("last_login", (self.current_user or {}).get("last_login", "-")))
        display_reg = str(profile.get("created_at", "-"))
        display_logins = str(profile.get("login_count", profile.get("logins", 0)))

        _safe_text("profile_banner_name", display_name)
        _safe_text("profile_banner_desc",
                   "系统管理员账号，可管理客流检测、数据统计与系统配置。" if display_role == "管理员"
                   else "员工账号，可查看客流检测与数据统计信息。")

        if display_role == "管理员":
            _safe_badge("profile_banner_role_badge", display_role, "#DBEAFE", "#1E40AF")
        else:
            _safe_badge("profile_banner_role_badge", display_role, "#FEF3C7", "#92400E")

        if display_status == "已启用":
            _safe_badge("profile_banner_status_badge", display_status, "#D1FAE5", "#065F46")
        else:
            _safe_badge("profile_banner_status_badge", display_status, "#FEE2E2", "#991B1B")

        _safe_text("profile_basic_username", display_name)
        _safe_text("profile_basic_role", display_role)
        _safe_text("profile_basic_status", display_status)
        _safe_text("profile_basic_reg_time", display_reg)
        _safe_text("profile_login_last", display_login)
        _safe_text("profile_login_count", display_logins)
        _safe_text("profile_sec_perm_level", display_role)
        _safe_text("profile_sec_pwd_status", "正常")
        _safe_text("profile_sec_protection", "已开启" if display_status == "已启用" else "未启用")

        if hasattr(self, "profile_avatar") and display_name:
            initial = display_name[0].upper() if display_name else "U"
            try:
                self._draw_avatar(initial)
            except Exception:
                pass

    def _open_change_password_dialog(self):
        username = (self.current_user or {}).get("username", "").strip()
        if not username:
            QMessageBox.warning(self, "修改密码", "当前没有可用账号信息。")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("修改密码")
        dialog_layout = QVBoxLayout(dialog)
        form = QFormLayout()
        old_edit = QLineEdit()
        old_edit.setEchoMode(QLineEdit.Password)
        new_edit = QLineEdit()
        new_edit.setEchoMode(QLineEdit.Password)
        confirm_edit = QLineEdit()
        confirm_edit.setEchoMode(QLineEdit.Password)
        form.addRow("旧密码", old_edit)
        form.addRow("新密码", new_edit)
        form.addRow("确认新密码", confirm_edit)
        dialog_layout.addLayout(form)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.button(QDialogButtonBox.Ok).setText("提交修改")
        button_box.button(QDialogButtonBox.Cancel).setText("取消")
        dialog_layout.addWidget(button_box)

        def submit_password_change():
            old_password = old_edit.text().strip()
            new_password = new_edit.text().strip()
            confirm_password = confirm_edit.text().strip()
            if not old_password or not new_password:
                QMessageBox.warning(dialog, "修改密码", "旧密码和新密码不能为空。")
                return
            if new_password != confirm_password:
                QMessageBox.warning(dialog, "修改密码", "两次输入的新密码不一致。")
                return
            success, message = self.auth_manager.change_password(username, old_password, new_password)
            if success:
                QMessageBox.information(dialog, "修改密码", message)
                dialog.accept()
                self._refresh_profile_page()
                return
            QMessageBox.warning(dialog, "修改密码", message)

        button_box.accepted.connect(submit_password_change)
        button_box.rejected.connect(dialog.reject)
        dialog.exec_()

    def _load_sample_employees(self):
        self._employee_rows = [
            {
                "id": 1001,
                "username": "admin",
                "fullname": "系统管理员",
                "role": "管理员",
                "status": "已启用",
                "created_at": "2024-01-10 09:00:00",
                "last_login": "2026-05-06 08:30:00",
            },
            {
                "id": 1002,
                "username": "zhangsan",
                "fullname": "张三",
                "role": "员工",
                "status": "已启用",
                "created_at": "2025-02-18 14:20:00",
                "last_login": "2026-05-05 17:05:00",
            },
            {
                "id": 1003,
                "username": "lisi",
                "fullname": "李四",
                "role": "员工",
                "status": "已禁用",
                "created_at": "2025-06-01 10:00:00",
                "last_login": "2026-04-28 12:10:00",
            },
        ]

    def _load_sample_employee_logs(self):
        self._employee_logs = [
            {"time": "2026-05-06 09:12:00", "operator": "admin", "action": "启用账号", "target": "zhangsan", "result": "成功"},
            {"time": "2026-05-06 09:08:00", "operator": "admin", "action": "重置密码", "target": "lisi", "result": "成功"},
            {"time": "2026-05-06 09:02:00", "operator": "admin", "action": "新增员工", "target": "wangwu", "result": "成功"},
        ]

    def refresh_employee_list(self):
        if not hasattr(self, "emp_table"):
            return
        rows = getattr(self, "_employee_rows", [])
        keyword = self.emp_search_input.text().strip().lower() if hasattr(self, "emp_search_input") else ""
        role_filter = self.emp_role_filter.currentText() if hasattr(self, "emp_role_filter") else "全部"
        status_filter = self.emp_status_filter.currentText() if hasattr(self, "emp_status_filter") else "全部"

        def match(row):
            if keyword and keyword not in str(row.get("username", "")).lower() and keyword not in str(row.get("fullname", "")).lower():
                return False
            if role_filter != "全部" and row.get("role") != role_filter:
                return False
            if status_filter != "全部" and row.get("status") != status_filter:
                return False
            return True

        filtered_rows = [row for row in rows if match(row)]
        self.emp_table.setRowCount(0)
        for row in filtered_rows:
            row_index = self.emp_table.rowCount()
            self.emp_table.insertRow(row_index)
            values = [
                row.get("id", ""),
                row.get("username", ""),
                row.get("fullname", ""),
                row.get("role", ""),
                row.get("status", ""),
                row.get("created_at", ""),
                row.get("last_login", ""),
            ]
            for col, value in enumerate(values):
                self.emp_table.setItem(row_index, col, QTableWidgetItem(str(value)))

            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(0, 0, 0, 0)
            action_layout.setSpacing(6)
            inspect_btn = QPushButton("查看")
            inspect_btn.clicked.connect(lambda _=False, rid=row.get("id"): QMessageBox.information(self, "员工信息", f"员工编号：{rid}"))
            action_layout.addWidget(inspect_btn)
            action_layout.addStretch()
            self.emp_table.setCellWidget(row_index, 7, action_widget)

    def refresh_employee_logs(self):
        if not hasattr(self, "employee_log_table"):
            return
        self.employee_log_table.setRowCount(0)
        for log in getattr(self, "_employee_logs", []):
            row_index = self.employee_log_table.rowCount()
            self.employee_log_table.insertRow(row_index)
            self.employee_log_table.setItem(row_index, 0, QTableWidgetItem(str(log.get("time", ""))))
            self.employee_log_table.setItem(row_index, 1, QTableWidgetItem(str(log.get("operator", ""))))
            self.employee_log_table.setItem(row_index, 2, QTableWidgetItem(str(log.get("action", ""))))
            self.employee_log_table.setItem(row_index, 3, QTableWidgetItem(str(log.get("target", ""))))
            self.employee_log_table.setItem(row_index, 4, QTableWidgetItem(str(log.get("result", ""))))

    def _append_employee_log(self, action_type: str, target_employee: str, result: str):
        log_item = {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "operator": (self.current_user or {}).get("username", "未知"),
            "action": action_type,
            "target": target_employee,
            "result": result,
        }
        self._employee_logs = [log_item] + getattr(self, "_employee_logs", [])[:19]
        self.refresh_employee_logs()

    def apply_employee_filters(self):
        self.refresh_employee_list()

    def reset_employee_filters(self):
        if hasattr(self, "emp_search_input"):
            self.emp_search_input.clear()
        if hasattr(self, "emp_role_filter"):
            self.emp_role_filter.setCurrentText("全部")
        if hasattr(self, "emp_status_filter"):
            self.emp_status_filter.setCurrentText("全部")
        self.refresh_employee_list()

    def _selected_employee_row(self):
        if not hasattr(self, "emp_table"):
            return None
        row_index = self.emp_table.currentRow()
        if row_index < 0:
            return None
        emp_id_item = self.emp_table.item(row_index, 0)
        if emp_id_item is None:
            return None
        emp_id = emp_id_item.text().strip()
        for row in getattr(self, "_employee_rows", []):
            if str(row.get("id")) == emp_id:
                return row
        return None

    def handle_add_employee(self):
        if not self._require_admin("新增员工"):
            return
        QMessageBox.information(self, "新增员工", "这里预留新增员工对话框接口，后续接数据库。")
        self._append_employee_log("新增员工", "新员工", "成功")

    def handle_edit_employee(self):
        if not self._require_admin("编辑员工"):
            return
        row = self._selected_employee_row()
        if row is None:
            QMessageBox.warning(self, "编辑员工", "请先选择一名员工。")
            return
        QMessageBox.information(self, "编辑员工", f"这里预留编辑接口：{row.get('username')}")
        self._append_employee_log("编辑员工", str(row.get("username", "")), "成功")

    def handle_delete_employee(self):
        if not self._require_admin("删除员工"):
            return
        row = self._selected_employee_row()
        if row is None:
            QMessageBox.warning(self, "删除员工", "请先选择一名员工。")
            return
        current_username = (self.current_user or {}).get("username", "")
        if row.get("username") == current_username and self._is_admin():
            QMessageBox.warning(self, "删除受限", "不能删除当前正在登录的管理员账号。")
            return
        confirm = QMessageBox.question(self, "确认删除", f"确认删除员工 {row.get('username')} 吗？", QMessageBox.Yes | QMessageBox.No)
        if confirm != QMessageBox.Yes:
            return
        self._employee_rows = [item for item in getattr(self, "_employee_rows", []) if item.get("id") != row.get("id")]
        self.refresh_employee_list()
        self._append_employee_log("删除员工", str(row.get("username", "")), "成功")

    def handle_reset_employee_password(self):
        if not self._require_admin("重置密码"):
            return
        row = self._selected_employee_row()
        if row is None:
            QMessageBox.warning(self, "重置密码", "请先选择一名员工。")
            return
        confirm = QMessageBox.question(self, "确认重置密码", f"确认重置员工 {row.get('username')} 的密码吗？", QMessageBox.Yes | QMessageBox.No)
        if confirm != QMessageBox.Yes:
            return
        QMessageBox.information(self, "重置密码", "已预留重置密码接口，后续接数据库与重置逻辑。")
        self._append_employee_log("重置密码", str(row.get("username", "")), "成功")

    def handle_enable_employee(self):
        if not self._require_admin("启用账号"):
            return
        row = self._selected_employee_row()
        if row is None:
            QMessageBox.warning(self, "启用账号", "请先选择一名员工。")
            return
        row["status"] = "已启用"
        self.refresh_employee_list()
        self._append_employee_log("启用账号", str(row.get("username", "")), "成功")

    def handle_disable_employee(self):
        if not self._require_admin("禁用账号"):
            return
        row = self._selected_employee_row()
        if row is None:
            QMessageBox.warning(self, "禁用账号", "请先选择一名员工。")
            return
        row["status"] = "已禁用"
        self.refresh_employee_list()
        self._append_employee_log("禁用账号", str(row.get("username", "")), "成功")

    def on_export_results_requested(self):
        if not self._check_permission("can_export_data", "导出结果摘要"):
            return
        default_dir = writable_path("outputs/gui_run")
        default_dir.mkdir(parents=True, exist_ok=True)
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出结果摘要",
            str(default_dir / "summary.txt"),
            "Text Files (*.txt);;All Files (*)",
        )
        if not file_path:
            return

        summary_lines = [
            f"用户: {(self.current_user or {}).get('username', '-')}",
            f"角色: {(self.current_user or {}).get('role', '-')}",
            f"总通行人数: {self.current_total_people}",
            f"当前批次: {self.current_run_id or '-'}",
            f"{self.stats_query_total.title_label.text()}: {self.stats_query_total.value_label.text()}",
            f"{self.stats_query_up.title_label.text()}: {self.stats_query_up.value_label.text()}",
            f"{self.stats_query_down.title_label.text()}: {self.stats_query_down.value_label.text()}",
            f"{self.stats_query_fps.title_label.text()}: {self.stats_query_fps.value_label.text()}",
            f"当前设备: {self.current_device}",
            f"平均 FPS: {self.current_avg_fps:.2f}",
            f"视频分辨率: {self.current_video_resolution}",
        ]
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("\n".join(summary_lines))
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"结果写入失败: {exc}")
            return

        self.log_panel.append_log(f"结果已导出: {file_path}")

    def export_filtered_events(self):
        if not self._check_permission("can_export_data", "导出事件表"):
            return
        default_dir = writable_path("outputs/gui_run")
        default_dir.mkdir(parents=True, exist_ok=True)
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出事件表",
            str(default_dir / "events_filtered.csv"),
            "CSV Files (*.csv)",
        )
        if not file_path:
            return
        try:
            self.filtered_event_table_panel.export_csv(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"事件表导出失败: {exc}")
            return
        self.log_panel.append_log(f"事件表已导出: {file_path}")

    def export_statistics_charts(self):
        if not self._check_permission("can_export_data", "导出图表截图"):
            return
        default_dir = writable_path("outputs/gui_run")
        default_dir.mkdir(parents=True, exist_ok=True)
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出图表截图",
            str(default_dir / "statistics_charts.png"),
            "PNG Image (*.png)",
        )
        if not file_path:
            return
        try:
            base = Path(file_path)
            self.stats_history_panel.figure.savefig(str(base), dpi=180)
            fps_path = base.with_name(base.stem + "_fps" + base.suffix)
            self.stats_fps_panel.figure.savefig(str(fps_path), dpi=180)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"图表导出失败: {exc}")
            return
        self.log_panel.append_log(f"图表截图已导出: {file_path}")

    def export_database_history_results(self):
        if not self._check_permission("can_export_data", "导出数据库历史结果"):
            return
        default_dir = writable_path("outputs/gui_run")
        default_dir.mkdir(parents=True, exist_ok=True)
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "导出数据库历史结果",
            str(default_dir / "database_history.xlsx"),
            "Excel Workbook (*.xlsx);;CSV Files (*.csv)",
        )
        if not file_path:
            return

        context = self._get_statistics_query_context()
        history_rows = context["history_rows"]
        events = context["events"]
        sessions = context["sessions"]
        fps_rows = context.get("fps_rows", [])

        try:
            if file_path.lower().endswith(".csv") or "CSV Files" in selected_filter:
                self._export_history_csv(file_path, context)
            else:
                self._export_history_xlsx(file_path, context)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"历史结果导出失败: {exc}")
            return

        self.log_panel.append_log(f"数据库历史结果已导出: {file_path}")

    def _load_replay_annotations(self, session_row: dict) -> tuple[list, list]:
        source_path = str(session_row.get("source_path") or "").strip()
        if source_path:
            try:
                db_ann = self.history_db.query_annotations(source_path)
                if db_ann and (db_ann.get("roi_points") or db_ann.get("line_points")):
                    return list(db_ann.get("roi_points", []) or []), list(db_ann.get("line_points", []) or [])[:2]
            except Exception:
                pass

        config_path = str(session_row.get("config_path") or "").strip()
        if not config_path:
            return [], []

        config_file = Path(config_path)
        if not config_file.exists():
            return [], []

        try:
            with config_file.open("r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f) or {}
        except Exception:
            return [], []

        counting_cfg = cfg.get("counting", {}) if isinstance(cfg, dict) else {}
        roi_cfg = counting_cfg.get("roi", {})
        line_cfg = counting_cfg.get("line", {})
        roi_points = roi_cfg.get("polygon", []) if isinstance(roi_cfg, dict) else []
        line_points = line_cfg.get("points", []) if isinstance(line_cfg, dict) else []
        return self._normalize_annotation_points_for_source(cfg, session_row.get("source_path"), roi_points or [], (line_points or [])[:2])

    def _probe_video_frame_size(self, video_path: str | None) -> tuple[int, int]:
        if not video_path:
            return 0, 0
        try:
            import cv2

            capture = cv2.VideoCapture(str(video_path))
            if not capture.isOpened():
                return 0, 0
            width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
            height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
            if width <= 0 or height <= 0:
                ok, frame = capture.read()
                if ok and frame is not None:
                    height, width = frame.shape[:2]
            capture.release()
            return max(0, int(width)), max(0, int(height))
        except Exception:
            return 0, 0

    def _annotation_frame_size_from_cfg(self, cfg: dict | None, source_w: int, source_h: int) -> tuple[int, int]:
        annotation_cfg = cfg.get("annotation_frame_size", {}) if isinstance(cfg, dict) else {}
        if isinstance(annotation_cfg, dict):
            width = int(annotation_cfg.get("width", 0) or 0)
            height = int(annotation_cfg.get("height", 0) or 0)
            if width > 0 and height > 0:
                return width, height

        performance_cfg = cfg.get("performance", {}) if isinstance(cfg, dict) else {}
        max_width = int(performance_cfg.get("max_width", 1280) or 1280)
        max_height = int(performance_cfg.get("max_height", 800) or 800)
        if source_w > 0 and source_h > 0:
            scale = min(1.0, min(max_width / float(source_w), max_height / float(source_h)))
            proc_w = max(1, int(round(source_w * scale)))
            proc_h = max(1, int(round(source_h * scale)))
            return proc_w, proc_h
        return 0, 0

    def _normalize_annotation_points_for_source(self, cfg: dict | None, source_path: str | None, roi_points: list, line_points: list) -> tuple[list, list]:
        source_w = source_h = 0
        if self.current_source_frame_size and all(int(v or 0) > 0 for v in self.current_source_frame_size):
            source_w, source_h = self.current_source_frame_size
        else:
            source_w, source_h = self._probe_video_frame_size(source_path)

        if source_w <= 0 or source_h <= 0:
            return list(roi_points or []), list(line_points or [])[:2]

        annotation_w, annotation_h = self._annotation_frame_size_from_cfg(cfg, source_w, source_h)
        coord_space = str(cfg.get("annotation_frame_size", {}).get("coord_space", "") if isinstance(cfg, dict) else "").lower()
        if coord_space != "annotation":
            return list(roi_points or []), list(line_points or [])[:2]

        if annotation_w > 0 and annotation_h > 0 and (annotation_w != source_w or annotation_h != source_h):
            roi_points = frame_points_to_frame_points(roi_points, annotation_w, annotation_h, source_w, source_h)
            line_points = frame_points_to_frame_points(line_points, annotation_w, annotation_h, source_w, source_h)
        return roi_points or [], (line_points or [])[:2]

    def _build_replay_geometry_scale(self, session_row: dict) -> tuple[float, float]:
        video_path = str(session_row.get("source_path") or "").strip()
        config_path = str(session_row.get("config_path") or "").strip()
        if not video_path:
            return 1.0, 1.0

        try:
            import cv2

            capture = cv2.VideoCapture(video_path)
            if not capture.isOpened():
                return 1.0, 1.0
            raw_width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
            raw_height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
            capture.release()
            if raw_width <= 0 or raw_height <= 0:
                return 1.0, 1.0
        except Exception:
            return 1.0, 1.0

        max_width, max_height = 1280, 800
        if config_path:
            try:
                with Path(config_path).open("r", encoding="utf-8") as f:
                    cfg = yaml.safe_load(f) or {}
                perf_cfg = cfg.get("performance", {}) if isinstance(cfg, dict) else {}
                max_width = int(perf_cfg.get("max_width", max_width) or max_width)
                max_height = int(perf_cfg.get("max_height", max_height) or max_height)
            except Exception:
                pass

        proc_width, proc_height = self._compute_adaptive_size(raw_width, raw_height, max_width=max_width, max_height=max_height)
        if proc_width <= 0 or proc_height <= 0:
            return 1.0, 1.0
        return raw_width / float(proc_width), raw_height / float(proc_height)

    def _scale_replay_point(self, point, scale_x: float, scale_y: float):
        if point is None:
            return None
        try:
            x, y = point
            return (float(x) * scale_x, float(y) * scale_y)
        except Exception:
            return None

    def _build_replay_payload(self, session_id: int):
        sessions = self.history_db.query_sessions(session_ids=[session_id], order="desc", limit=1)
        if not sessions:
            return None

        session_row = sessions[0]
        video_path = str(session_row.get("source_path") or "").strip()
        if not video_path:
            return None

        source_w, source_h = self._probe_video_frame_size(video_path)
        if source_w <= 0 or source_h <= 0:
            source_w, source_h = 0, 0
        cfg = None
        config_path = str(session_row.get("config_path") or "").strip()
        if config_path:
            try:
                with Path(config_path).open("r", encoding="utf-8") as f:
                    cfg = yaml.safe_load(f) or {}
            except Exception:
                cfg = None
        annotation_w, annotation_h = self._annotation_frame_size_from_cfg(cfg, source_w, source_h)

        payload = {
            "session_row": session_row,
            "video_path": video_path,
            "trajectory_rows": self.history_db.query_trajectory_points(session_ids=[session_id]),
            "traffic_rows": self.history_db.query_traffic_samples(session_ids=[session_id]),
            "fps_rows": self.history_db.query_fps_history(session_ids=[session_id]),
            "event_rows": self.history_db.query_events(session_ids=[session_id]),
            "source_frame_size": (source_w, source_h),
            "annotation_frame_size": (annotation_w, annotation_h),
        }
        roi_points, line_points = self._load_replay_annotations(session_row)
        payload["roi_points"] = [tuple(map(float, point)) for point in roi_points if point is not None]
        payload["line_points"] = [tuple(map(float, point)) for point in line_points if point is not None]

        coord_space = str(cfg.get("annotation_frame_size", {}).get("coord_space", "") if isinstance(cfg, dict) else "").lower()
        if coord_space == "annotation" and source_w > 0 and source_h > 0 and annotation_w > 0 and annotation_h > 0 and (annotation_w != source_w or annotation_h != source_h):
            normalized_trajectory_rows = []
            for row in payload["trajectory_rows"]:
                if not isinstance(row, dict):
                    continue
                normalized_row = dict(row)
                point_keys = [("x1", "y1"), ("x2", "y2"), ("cx", "cy")]
                for x_key, y_key in point_keys:
                    if normalized_row.get(x_key) is None or normalized_row.get(y_key) is None:
                        continue
                    point = frame_points_to_frame_points(
                        [(float(normalized_row.get(x_key)), float(normalized_row.get(y_key)))],
                        annotation_w,
                        annotation_h,
                        source_w,
                        source_h,
                    )
                    if point:
                        normalized_row[x_key], normalized_row[y_key] = float(point[0][0]), float(point[0][1])
                normalized_trajectory_rows.append(normalized_row)
            payload["trajectory_rows"] = normalized_trajectory_rows
        return payload

    def open_selected_batch_replay(self):
        context = self._get_statistics_query_context()
        session_row = self._get_selected_statistics_run()
        if not isinstance(session_row, dict) or int(session_row.get("session_id") or session_row.get("id") or 0) <= 0:
            session_row = context.get("selected_run")
        if not session_row:
            QMessageBox.information(self, "轨迹回放", "当前没有可回放的检测批次。")
            return

        session_id = int(session_row.get("id") or session_row.get("session_id") or 0)
        if session_id <= 0:
            QMessageBox.information(self, "轨迹回放", "当前选中的批次缺少有效 ID。")
            return

        payload = self._build_replay_payload(session_id)
        if payload is None:
            QMessageBox.warning(self, "轨迹回放", "无法加载当前批次的回放数据。")
            return

        try:
            dialog = TrajectoryReplayDialog(
                session_row=payload["session_row"],
                video_path=payload["video_path"],
                trajectory_rows=payload["trajectory_rows"],
                traffic_rows=payload["traffic_rows"],
                fps_rows=payload["fps_rows"],
                event_rows=payload["event_rows"],
                roi_points=payload.get("roi_points", []),
                line_points=payload.get("line_points", []),
                parent=self,
            )
        except Exception as exc:
            QMessageBox.critical(self, "轨迹回放", f"打开回放失败: {exc}")
            return

        self._trajectory_replay_dialog = dialog
        dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def _export_history_csv(self, file_path: str, context: dict):
        history_rows = context["history_rows"]
        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["item", "value"])
            writer.writerow(["mode", context.get("mode", "")])
            writer.writerow(["selected_run", context.get("selected_run_label", "")])
            writer.writerow([])
            writer.writerow(["bucket_start", "bucket_end", "source_name", "up", "down", "total", "records", "fps", "avg_fps"])
            for row in history_rows:
                writer.writerow([
                    row.get("bucket_start", ""),
                    row.get("bucket_end", ""),
                    row.get("source_name", ""),
                    row.get("up", 0),
                    row.get("down", 0),
                    row.get("total", 0),
                    row.get("records", 0),
                    f"{float(row.get('fps') or 0.0):.2f}",
                    f"{float(row.get('avg_fps') or 0.0):.2f}",
                ])

    def _export_history_xlsx(self, file_path: str, context: dict):
        sessions = context["sessions"]
        history_rows = context["history_rows"]
        events = context["events"]
        fps_rows = context.get("fps_rows", [])
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["item", "value"])
        summary_sheet.append(["mode", context.get("mode", "")])
        summary_sheet.append(["start_time", self.query_time_start.text().strip() or "全部"])
        summary_sheet.append(["end_time", self.query_time_end.text().strip() or "全部"])
        summary_sheet.append(["video", self.query_video_name.currentText()])
        summary_sheet.append(["run_batch", context.get("selected_run_label", "")])
        summary_sheet.append(["direction", self.query_direction.currentText()])
        summary_sheet.append(["bucket", self.query_bucket.currentText()])
        summary_sheet.append(["history_rows", len(history_rows)])
        summary_sheet.append(["events", len(events)])
        summary_sheet.append(["sessions", len(sessions)])
        summary_sheet.append([])
        summary_sheet.append(["bucket_start", "bucket_end", "source_name", "up", "down", "total", "records"])
        for row in history_rows:
            summary_sheet.append([
                row.get("bucket_start", ""),
                row.get("bucket_end", ""),
                row.get("source_name", ""),
                row.get("up", 0),
                row.get("down", 0),
                row.get("total", 0),
                row.get("records", 0),
            ])

        sessions_sheet = workbook.create_sheet("Sessions")
        sessions_sheet.append(["id", "run_id", "source_name", "started_at", "ended_at", "conf", "iou", "avg_fps", "up_count", "down_count", "total_count", "detector_type", "config_path"])
        for row in sessions:
            sessions_sheet.append([
                row.get("id", ""),
                row.get("run_id", ""),
                row.get("source_name", ""),
                row.get("started_at", ""),
                row.get("ended_at", ""),
                row.get("conf", ""),
                row.get("iou", ""),
                row.get("avg_fps", ""),
                row.get("up_count", ""),
                row.get("down_count", ""),
                row.get("total_count", ""),
                row.get("detector_type", ""),
                row.get("config_path", ""),
            ])

        events_sheet = workbook.create_sheet("Events")
        events_sheet.append(["timestamp", "event_type", "direction", "target", "track_id", "value", "source_name", "run_id", "detect_time", "frame_idx"])
        for row in events:
            events_sheet.append([
                row.get("timestamp", ""),
                row.get("event_type", ""),
                row.get("direction", ""),
                row.get("target", ""),
                row.get("track_id", ""),
                row.get("value", ""),
                row.get("source_name", ""),
                row.get("run_id", ""),
                row.get("detect_time", ""),
                row.get("frame_idx", ""),
            ])

        fps_sheet = workbook.create_sheet("FPS")
        fps_sheet.append(["timestamp", "fps", "avg_fps", "source_name", "run_id", "detect_time"])
        for row in fps_rows:
            fps_sheet.append([
                row.get("timestamp", ""),
                row.get("fps", ""),
                row.get("avg_fps", ""),
                row.get("source_name", ""),
                row.get("run_id", ""),
                row.get("detect_time", ""),
            ])

        workbook.save(file_path)

    def start_processing(self):
        video_path = self.control_panel.get_video_path()
        if not video_path:
            self.log_panel.append_log("请先选择视频源")
            return

        runtime_params = self.annotation_panel.get_params()
        detector_cfg = self.current_config.setdefault("detector", {})
        detector_cfg["conf"] = float(runtime_params.get("conf", detector_cfg.get("conf", 0.5)))
        detector_cfg["iou"] = float(runtime_params.get("iou", detector_cfg.get("iou", 0.45)))
        active_model = get_active_model(self.model_registry) if isinstance(self.model_registry, dict) else None
        if active_model is not None:
            detector_cfg["model_path"] = active_model.get("path", detector_cfg.get("model_path", ""))
            if active_model.get("imgsz") is not None:
                detector_cfg["imgsz"] = int(active_model["imgsz"])
            if active_model.get("device") is not None:
                detector_cfg["device"] = active_model["device"]
        self._sync_model_registry_from_current_config(save=True)

        self.current_runtime_start = time.time()
        self.current_frame_number = 0
        self.current_total_frames = 0
        self.current_total_people = 0
        self.current_up_count = 0
        self.current_down_count = 0
        self.current_detecting = True
        self.current_avg_fps = 0.0
        self.current_run_id = None
        self.current_session_id = None
        self._active_alert_keys = set()
        self._alert_last_emit_at = {}
        self._last_realtime_stats = None
        self.home_metric_total.set_value("0", "本次会话累计")
        self.home_metric_up.set_value("0", "上行人数")
        self.home_metric_down.set_value("0", "下行人数")
        self.home_metric_current.set_value("0", "当前画面")
        self.home_metric_fps.set_value("0.0", "实时性能")
        self.detect_metric_total.set_value("0", "实时统计")
        self.detect_metric_up.set_value("0", "上行人数")
        self.detect_metric_down.set_value("0", "下行人数")
        self.detect_metric_current.set_value("0", "当前画面")
        self.detect_metric_fps.set_value("0.0", "实时性能")
        self.worker.set_config(self.current_config)
        self.worker.set_source(video_path)
        self.worker.set_runtime_params(runtime_params)
        self.worker.set_annotations(self.video_panel._roi_points, self.video_panel._line_points)

        ensure_torch_preloaded()
        self.current_device = self._device_label()

        self.home_trend_panel.reset()
        self.stats_history_panel.reset()
        self.stats_fps_panel.reset()
        self.event_table_panel.reset()
        self.filtered_event_table_panel.reset()
        self.clear_alert_history(silent=True)
        self.home_summary_list.clear()
        self.event_history = []
        self.trend_history = []
        self.fps_history = []
        self._refresh_system_info()
        self._refresh_home_status_card()
        self._update_alert_status_label()

        self.worker.start()
        self._stale_timer.start()
        self._last_frame_position = 0
        self.video_player.set_position(0, 0)
        self.video_player.set_playing(True)
        self.control_panel.btn_start.setEnabled(False)
        self.control_panel.btn_stop.setEnabled(True)
        self.control_panel.btn_pause.setEnabled(True)
        self.control_panel.btn_pause.setText("暂停")
        self.log_panel.append_log("线程启动")
        if hasattr(self, "video_placeholder"):
            self.video_placeholder.hide()
            self.video_player.show()

    def stop_processing(self):
        self.worker.stop()
        self._stale_timer.stop()
        self.log_panel.append_log("正在停止")
        if hasattr(self, "video_placeholder"):
            self.video_placeholder.show()
            self.video_player.hide()

    def toggle_pause(self):
        if not self.worker.isRunning():
            return
        if self.worker.paused:
            self.worker.resume()
            self.control_panel.btn_pause.setText("暂停")
            self.log_panel.append_log("继续处理")
        else:
            self.worker.pause()
            self.control_panel.btn_pause.setText("继续")
            self.log_panel.append_log("已暂停")

    def _reconnect_worker_signals(self):
        self.worker.new_frame.connect(self.video_player.update_frame)
        self.worker.stats_updated.connect(self.on_stats_updated)
        self.worker.trend_updated.connect(self.on_trend_updated)
        self.worker.event_emitted.connect(self.on_event_emitted)
        self.worker.log_message.connect(self.log_panel.append_log)
        self.worker.finished.connect(self.on_process_finished)
        self.worker.frame_position.connect(self._on_frame_position)
        self.worker.play_state_changed.connect(self.video_player.set_playing)

    def _on_frame_position(self, current: int, total: int):
        self._last_frame_position = current
        self.video_player.set_position(current, total)

    def _on_seek_requested(self, frame_number: int):
        if self.worker.isRunning():
            self.worker.seek_to_frame(frame_number)

    def _on_step_forward(self, delta: int):
        target = self._last_frame_position + delta
        self._on_seek_requested(target)

    def _on_step_backward(self, delta: int):
        target = max(0, self._last_frame_position - delta)
        self._on_seek_requested(target)

    def on_process_finished(self):
        self.control_panel.btn_start.setEnabled(True)
        self.control_panel.btn_stop.setEnabled(False)
        self.control_panel.btn_pause.setEnabled(False)
        self.control_panel.btn_pause.setText("暂停")
        self.current_runtime_start = None
        self.current_detecting = False
        self._stale_timer.stop()
        self._active_alert_keys = set()
        self._update_alert_status_label()
        self.video_player.set_playing(False)
        self._last_frame_position = 0
        if hasattr(self, "video_placeholder"):
            self.video_placeholder.show()
            self.video_player.hide()
        try:
            self.history_db.close()
        except Exception:
            pass
        self.history_db = DatabaseManager(writable_path("outputs/traffic.db"))

        self._auto_refresh_statistics_page()

        self._refresh_home_status_card()
        self._refresh_system_info()

        self.worker = WorkerThread()
        self.worker.set_config(self.current_config)
        self._reconnect_worker_signals()

    def _auto_refresh_statistics_page(self):
        self.query_time_start.clear()
        self.query_time_end.clear()
        self.query_stats_mode.setCurrentText("最新一次检测")
        self.query_direction.setCurrentText("All")
        self.query_bucket.setCurrentText("按分钟")
        self._refresh_statistics_sources()
        self.query_video_name.setCurrentIndex(0)
        self._refresh_statistics_run_choices()
        self.apply_statistics_filters()

    def on_trend_updated(self, payload):
        if bool(payload.get("reset", False)):
            self.home_trend_panel.reset()
            self.stats_history_panel.reset()
            return
        self.trend_history.append(
            {
                "minute": int(payload.get("minute", 0)),
                "up": int(payload.get("up", 0)),
                "down": int(payload.get("down", 0)),
                "total": int(payload.get("total", 0)),
            }
        )
        self.stats_history_panel.update_trend(payload)

    def on_stats_updated(self, stats):
        self.stats_panel.update_stats(stats)
        self._last_realtime_stats = dict(stats)
        session_id = stats.get("session_id")
        run_id = stats.get("run_id")
        if session_id is not None:
            try:
                self.current_session_id = int(session_id)
            except Exception:
                pass
        if run_id is not None:
            self.current_run_id = str(run_id)
        self.home_metric_total.set_value(str(int(stats.get("total", 0))), "本次会话累计")
        self.home_metric_up.set_value(str(int(stats.get("up", 0))), "上行人数")
        self.home_metric_down.set_value(str(int(stats.get("down", 0))), "下行人数")
        self.home_metric_current.set_value(str(int(stats.get("current", 0))), "当前画面")
        self.home_metric_fps.set_value(f"{float(stats.get('fps', 0.0)):.1f}", "实时性能")

        self.detect_metric_total.set_value(str(int(stats.get("total", 0))), "实时统计")
        self.detect_metric_up.set_value(str(int(stats.get("up", 0))), "上行人数")
        self.detect_metric_down.set_value(str(int(stats.get("down", 0))), "下行人数")
        self.detect_metric_current.set_value(str(int(stats.get("current", 0))), "当前画面")
        self.detect_metric_fps.set_value(f"{float(stats.get('fps', 0.0)):.1f}", "实时性能")

        self.stats_fps_panel.update_fps(stats.get("fps", 0.0))
        self.fps_history.append(float(stats.get("fps", 0.0)))
        self.current_source_fps = float(stats.get("source_fps", self.current_source_fps))
        self.current_frame_number = int(stats.get("current_frame", self.current_frame_number))
        self.current_total_frames = int(stats.get("total_frames", self.current_total_frames))
        self.current_total_people = int(stats.get("total", self.current_total_people))
        self.current_up_count = int(stats.get("up", self.current_up_count))
        self.current_down_count = int(stats.get("down", self.current_down_count))
        self.current_avg_fps = float(stats.get("avg_fps", self.current_avg_fps))
        self._evaluate_realtime_alerts(stats)
        self._refresh_system_info(stats)

        cumulative_total = int(stats.get("total", 0))
        fps = float(stats.get("source_fps", self.current_source_fps or 25.0))
        current_frame = int(stats.get("current_frame", self.current_frame_number))
        current_minute = int((current_frame / fps) // 60) if fps > 0 else 0
        self.home_trend_panel.add_cumulative_total(cumulative_total, step_identifier=current_minute)

    def _check_stale_data(self):
        if self.current_detecting:
            self.home_trend_panel.mark_stale()

    def on_event_emitted(self, event):
        if bool(event.get("reset", False)):
            self.event_table_panel.reset()
            self.filtered_event_table_panel.reset()
            self.home_summary_list.clear()
            return
        self.event_table_panel.add_event_record(event, fps=self.current_source_fps)
        session_id = event.get("session_id")
        run_id = event.get("run_id")
        if session_id is not None:
            try:
                self.current_session_id = int(session_id)
            except Exception:
                pass
        if run_id is not None:
            self.current_run_id = str(run_id)
        event_name = str(event.get("value") or event.get("event_type") or "事件")
        track_id = event.get("track_id", "-")
        target = str(event.get("target", "-"))
        frame_idx = int(event.get("frame_idx", 0))
        sec = frame_idx / self.current_source_fps if self.current_source_fps > 1e-6 else 0.0
        mm = int(sec // 60)
        ss = int(sec % 60)
        direction = self._event_direction(event_name)
        event_record = {
            "timestamp": f"{mm:02d}:{ss:02d}",
            "seconds": int(sec),
            "event": event_name,
            "direction": direction,
            "track_id": str(track_id),
            "target": target,
            "video": self.current_source_name,
            "run_id": self.current_run_id or "",
        }
        self.event_history.append(event_record)
        self.home_summary_list.insertItem(0, QListWidgetItem(f"{mm:02d}:{ss:02d} · {event_name} · ID {track_id} · {target}"))
        while self.home_summary_list.count() > 5:
            self.home_summary_list.takeItem(self.home_summary_list.count() - 1)

    def on_video_selected(self, path):
        self.log_panel.append_log(f"已选择视频源: {path}")
        self.current_source_name = "摄像头 0" if str(path) == "0" else (Path(path).name if path else "未选择")
        if self.query_video_name.findText(self.current_source_name) < 0:
            self.query_video_name.addItem(self.current_source_name)
        self._refresh_statistics_run_choices()
        self._refresh_home_status_card()
        if path and str(path) != "0":
            import cv2
            from PyQt5.QtGui import QImage
            cap = cv2.VideoCapture(str(path))
            if cap.isOpened():
                width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                if width > 0 and height > 0:
                    self.current_video_resolution = f"{width} x {height}"
                    self.current_source_frame_size = (width, height)
                    self.video_panel.set_source_frame_size(width, height)
                    self._refresh_system_info()
                ret, frame = cap.read()
                if ret:
                    target_size = self._compute_adaptive_size(frame.shape[1], frame.shape[0])
                    if frame.shape[1] != target_size[0] or frame.shape[0] != target_size[1]:
                        frame = cv2.resize(frame, target_size)
                    rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    h, w, ch = rgb_image.shape
                    bytes_per_line = ch * w
                    # Keep a reference to rgb_image.data so it doesn't get garbage collected
                    # Or use image.copy()
                    q_img = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888).copy()
                    self.video_panel.update_frame(q_img)
                    if hasattr(self, "video_placeholder"):
                        self.video_placeholder.hide()
                        self.video_player.show()
                cap.release()

    def use_camera_source(self):
        self.control_panel.set_video_path("0")
        self.log_panel.append_log("视频源切换为摄像头 0")
        self.current_source_name = "摄像头 0"
        if self.query_video_name.findText(self.current_source_name) < 0:
            self.query_video_name.addItem(self.current_source_name)
        self._refresh_statistics_run_choices()
        self._refresh_home_status_card()
        import cv2
        from PyQt5.QtGui import QImage
        cap = cv2.VideoCapture(0)
        if cap.isOpened():
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            if width > 0 and height > 0:
                self.current_video_resolution = f"{width} x {height}"
                self.current_source_frame_size = (width, height)
                self.video_panel.set_source_frame_size(width, height)
                self._refresh_system_info()
            ret, frame = cap.read()
            if ret:
                target_size = self._compute_adaptive_size(frame.shape[1], frame.shape[0])
                if frame.shape[1] != target_size[0] or frame.shape[0] != target_size[1]:
                    frame = cv2.resize(frame, target_size)
                rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = rgb_image.shape
                bytes_per_line = ch * w
                q_img = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888).copy()
                self.video_panel.update_frame(q_img)
                if hasattr(self, "video_placeholder"):
                    self.video_placeholder.hide()
                    self.video_player.show()
            cap.release()

    def _compute_adaptive_size(self, orig_w, orig_h, max_width=1280, max_height=800):
        if orig_w <= 0 or orig_h <= 0:
            return max_width, max_height
        scale = min(1.0, min(max_width / orig_w, max_height / orig_h))
        new_w = max(1, int(orig_w * scale))
        new_h = max(1, int(orig_h * scale))
        return new_w, new_h

    def apply_runtime_params(self, params):
        self.worker.set_runtime_params(params)
        self.video_panel.set_show_flags(
            show_roi=params.get("show_roi", True),
            show_line=params.get("show_line", True),
        )

    def on_annotations_changed(self, _):
        self.worker.set_annotations(self.video_panel._roi_points, self.video_panel._line_points)
        self.annotation_panel.update_annotation_coords(self.video_panel._roi_points, self.video_panel._line_points)
        video_path = self.control_panel.get_video_path()
        if video_path and video_path != "0":
            try:
                self.history_db.upsert_annotations(video_path, self.video_panel._roi_points, self.video_panel._line_points)
            except Exception:
                pass

    def on_save_frame_requested(self):
        default_dir = writable_path("outputs/gui_run")
        default_dir.mkdir(parents=True, exist_ok=True)
        default_name = f"frame_{datetime.now():%Y%m%d_%H%M%S}.png"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存当前画面",
            str(default_dir / default_name),
            "PNG Image (*.png);;JPEG Image (*.jpg *.jpeg)",
        )
        if not file_path:
            return

        if self.video_panel.save_current_frame(file_path):
            self.log_panel.append_log(f"当前画面已保存: {file_path}")
        else:
            QMessageBox.warning(self, "保存失败", "当前没有可保存的画面")

    def load_yaml_config(self):
        if not self._require_admin("加载配置"):
            return
        default_dir = str(writable_path("config"))
        path, _ = QFileDialog.getOpenFileName(self, "加载YAML配置", default_dir, "YAML Files (*.yaml *.yml)")
        if not path:
            return

        try:
            with open(path, "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f) or {}
        except Exception as exc:
            QMessageBox.critical(self, "加载失败", f"配置读取失败: {exc}")
            return

        cfg["alert"] = self._normalize_alert_config(cfg.get("alert", {}))

        self.current_config_path = path
        self.current_config = cfg
        self.worker.set_config(cfg)
        self._apply_alert_config_to_ui(cfg.get("alert", self._default_alert_config()))
        self.current_detector_type = self._detector_type_label()
        self.current_onnx_enabled = bool(cfg.get("onnx", {}).get("enabled", False))
        self._sync_model_registry_from_current_config(save=True)
        self._refresh_system_info()
        self._refresh_home_status_card()

        src = str(cfg.get("source", ""))
        if src:
            self.control_panel.set_video_path(src)

        detector_cfg = cfg.get("detector", {})
        vis_cfg = cfg.get("visualization", {})
        privacy_cfg = cfg.get("privacy", {})
        face_blur_cfg = privacy_cfg.get("face_blur", {})
        counting_cfg = cfg.get("counting", {})
        line_cfg = counting_cfg.get("line", {})
        roi_cfg = counting_cfg.get("roi", {})

        source_path = str(cfg.get("source", "") or "")
        source_w, source_h = self._probe_video_frame_size(source_path)
        if source_w > 0 and source_h > 0:
            self.current_source_frame_size = (source_w, source_h)
            self.video_panel.set_source_frame_size(source_w, source_h)

        params = {
            "conf": float(detector_cfg.get("conf", 0.5)),
            "iou": float(detector_cfg.get("iou", 0.45)),
            "show_trail": bool(vis_cfg.get("show_trail", True)),
            "show_roi": bool(vis_cfg.get("draw_roi", True)),
            "show_line": bool(vis_cfg.get("draw_line", True)),
            "show_heatmap": bool(vis_cfg.get("show_heatmap", True)),
            "face_blur_enabled": bool(face_blur_cfg.get("enabled", False)),
            "draw_count_points": bool(cfg.get("debug", {}).get("draw_count_points", True)),
        }
        self.annotation_panel.set_params(params)

        roi_points, line_points = self._normalize_annotation_points_for_source(cfg, source_path, roi_cfg.get("polygon", []), line_cfg.get("points", []))
        cfg.setdefault("annotation_frame_size", {})
        if source_w > 0 and source_h > 0:
            cfg["annotation_frame_size"] = {"width": source_w, "height": source_h, "coord_space": "source"}
        self.video_panel.set_annotations(roi_points=roi_points, line_points=line_points)

        if source_path and source_path != "0":
            try:
                self.history_db.upsert_annotations(source_path, roi_points, line_points)
            except Exception:
                pass

        self.log_panel.append_log(f"配置已加载: {path}")
        self._append_management_log("配置管理", f"导入配置: {Path(path).name}", "成功")
        self._refresh_management_page()

    def save_yaml_config(self):
        if not self._require_admin("保存配置"):
            return
        default_dir = str(writable_path("config"))
        default_name = "pedestrian_gui_saved.yaml"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "保存YAML配置",
            str(Path(default_dir) / default_name),
            "YAML Files (*.yaml *.yml)",
        )
        if not path:
            return

        cfg = dict(self.current_config) if isinstance(self.current_config, dict) else self._default_config()
        params = self.annotation_panel.get_params()

        cfg["source"] = self.control_panel.get_video_path() or cfg.get("source", "")

        detector_cfg = cfg.setdefault("detector", {})
        detector_cfg["conf"] = float(params["conf"])
        detector_cfg["iou"] = float(params["iou"])

        vis_cfg = cfg.setdefault("visualization", {})
        vis_cfg["show_trail"] = bool(params["show_trail"])
        vis_cfg["draw_roi"] = bool(params["show_roi"])
        vis_cfg["draw_line"] = bool(params["show_line"])
        vis_cfg["show_heatmap"] = bool(params.get("show_heatmap", True))

        privacy_cfg = cfg.setdefault("privacy", {})
        face_blur_cfg = privacy_cfg.setdefault("face_blur", {})
        face_blur_cfg["enabled"] = bool(params.get("face_blur_enabled", False))

        cfg["alert"] = self._collect_alert_config_from_ui()

        debug_cfg = cfg.setdefault("debug", {})
        debug_cfg["draw_count_points"] = bool(params.get("draw_count_points", True))

        counting_cfg = cfg.setdefault("counting", {})
        roi_cfg = counting_cfg.setdefault("roi", {})
        line_cfg = counting_cfg.setdefault("line", {})
        roi_cfg["polygon"] = [list(map(int, p)) for p in self.video_panel._roi_points]
        line_cfg["points"] = [list(map(int, p)) for p in self.video_panel._line_points]

        if self.current_source_frame_size and all(int(v or 0) > 0 for v in self.current_source_frame_size):
            cfg["annotation_frame_size"] = {
                "width": int(self.current_source_frame_size[0]),
                "height": int(self.current_source_frame_size[1]),
                "coord_space": "source",
            }

        try:
            with open(path, "w", encoding="utf-8") as f:
                yaml.safe_dump(cfg, f, allow_unicode=True, sort_keys=False)
        except Exception as exc:
            QMessageBox.critical(self, "保存失败", f"配置写入失败: {exc}")
            return

        self.current_config_path = path
        self.current_config = cfg
        self.worker.set_config(cfg)
        self.current_detector_type = self._detector_type_label()
        self.current_onnx_enabled = bool(cfg.get("onnx", {}).get("enabled", False))
        video_path = self.control_panel.get_video_path()
        if video_path and video_path != "0":
            try:
                self.history_db.upsert_annotations(video_path, self.video_panel._roi_points, self.video_panel._line_points)
            except Exception:
                pass
        self._refresh_system_info()
        self.log_panel.append_log(f"配置已保存: {path}")
        self._append_management_log("配置管理", f"保存配置: {Path(path).name}", "成功")
        self._refresh_management_page()
        self._refresh_home_status_card()

    def show_about(self):
        QMessageBox.information(
            self,
            "关于",
            "基于YOLO的视频行人流量统计分析系统\n",
        )

    def _default_config(self):
        return {
            "source": "",
            "performance": {
                # UI 刷新/曲线重绘节流（不影响检测/计数逻辑）
                "ui_max_fps": 15.0,
                "stats_hz": 10.0,
                "trend_hz": 2.0,
                # 处理分辨率上限（越小越快，精度可能略降）
                "max_width": 1280,
                "max_height": 800,
            },
            "detector": {
                "model_path": str(resource_path("runs/train/yolov8m_ema_p3_896_sgd_cloud/weights/best.pt")),
                "conf": 0.5,
                "iou": 0.45,
            },
            "visualization": {
                "show_trail": True,
                "draw_roi": True,
                "draw_line": True,
                "show_heatmap": True,
                "heatmap_alpha": 0.35,
                "heatmap_sigma": 40,
                "heatmap_interval": 5,
            },
            "privacy": {
                "face_blur": {
                    "enabled": False,
                    "model_path": "models/yolov8n-face.pt",
                    "conf": 0.4,
                    "blur_type": "gaussian",
                    "blur_kernel": 25,
                }
            },
            "counting": {
                "roi": {"enabled": True, "polygon": []},
                "line": {"enabled": True, "points": []},
            },
            "alert": self._default_alert_config(),
            "debug": {"draw_count_points": True},
        }

    def _device_label(self):
        try:
            import torch
        except Exception:
            return "CPU"

        try:
            if torch.cuda.is_available():
                return f"GPU ({torch.cuda.get_device_name(0)})"
        except Exception:
            return "CPU"
        return "CPU"

    def _detector_type_label(self):
        detector_cfg = self.current_config.get("detector", {}) if isinstance(self.current_config, dict) else {}
        model_path = str(detector_cfg.get("model_path", ""))
        if not model_path:
            return "YOLO"

        path_obj = Path(model_path)
        model_dir = path_obj.parent.parent.name if len(path_obj.parts) >= 2 else path_obj.stem
        pretty = model_dir.replace("_", " ").strip()
        if pretty.lower().startswith("yolov8"):
            return pretty.replace("yolov8", "YOLOv8", 1)
        return pretty or "YOLO"

    def _format_runtime(self, seconds_value: float):
        seconds_value = max(0.0, float(seconds_value))
        total_seconds = int(seconds_value)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        if hours > 0:
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        return f"{minutes:02d}:{seconds:02d}"

    def _refresh_system_info(self, stats: dict | None = None):
        self.current_device = self._device_label()
        self.current_detector_type = self._detector_type_label()
        runtime_seconds = 0.0
        if self.current_runtime_start is not None:
            runtime_seconds = time.time() - self.current_runtime_start

        if stats is not None:
            self.current_source_fps = float(stats.get("source_fps", self.current_source_fps))

        info = {
            "model_path": self.current_config.get("detector", {}).get("model_path", "-"),
            "video_resolution": self.current_video_resolution,
            "device": self.current_device,
            "current_frame": self.current_frame_number,
            "total_frames": self.current_total_frames if self.current_total_frames > 0 else "未知",
            "total_people": self.current_total_people,
            "runtime": self._format_runtime(runtime_seconds),
            "avg_fps": f"{self.current_avg_fps:.2f}",
            "detector_type": self.current_detector_type,
            "tracker_type": self.current_tracker_type,
            "onnx_enabled": "是" if self.current_onnx_enabled else "否",
        }
        self.system_info_panel.update_info(info)

    def _start_tool_runner(self, cmd: list[str], cwd: str | None = None):
        runner = ToolRunnerThread(cmd, cwd=cwd)
        runner.log.connect(self.log_panel.append_log)
        runner.finished.connect(lambda rc, ref=runner: self._on_tool_finished(ref, rc))
        self.tool_runners.append(runner)
        runner.start()

    def _on_tool_finished(self, runner: ToolRunnerThread, return_code: int):
        self.log_panel.append_log(f"工具退出，返回码: {return_code}")
        if runner in self.tool_runners:
            self.tool_runners.remove(runner)
        runner.deleteLater()

    def _show_tool_dialog(self, title: str, fields: list[tuple[str, QWidget]]):
        dialog = ToolOptionsDialog(title, fields, self)
        result = dialog.exec_()
        if result != QDialog.Accepted:
            return None
        return dialog

    def on_export_onnx_requested(self):
        if not self._check_permission("can_switch_model", "导出 ONNX"):
            return
        weights, _ = QFileDialog.getOpenFileName(self, "选择权重文件 (.pt)", "", "PyTorch Weights (*.pt *.pth)")
        if not weights:
            return
        out_dir = QFileDialog.getExistingDirectory(self, "选择导出目录", str(Path(weights).parent))
        if not out_dir:
            return

        weights_path = weights
        out_path = out_dir

        imgsz = QSpinBox()
        imgsz.setRange(32, 4096)
        imgsz.setSingleStep(32)
        imgsz.setValue(800)

        opset = QSpinBox()
        opset.setRange(7, 25)
        opset.setValue(17)

        simplify = QCheckBox("启用 ONNX 简化")
        simplify.setChecked(True)

        dynamic = QCheckBox("动态 batch/shape")
        dynamic.setChecked(False)

        half = QCheckBox("导出 FP16")
        half.setChecked(False)

        dialog = self._show_tool_dialog(
            "ONNX 导出选项",
            [
                ("imgsz", imgsz),
                ("opset", opset),
                ("simplify", simplify),
                ("dynamic", dynamic),
                ("half", half),
            ],
        )
        if dialog is None:
            return

        export_imgsz = dialog.value("imgsz")
        export_opset = dialog.value("opset")
        export_simplify = dialog.value("simplify")
        export_dynamic = dialog.value("dynamic")
        export_half = dialog.value("half")

        expected_onnx = str(Path(out_path) / (Path(weights_path).stem + ".onnx"))

        self.export_thread = InlineExportThread(
            weights_path, out_path,
            export_imgsz, export_opset, export_simplify, export_dynamic, export_half,
        )
        self.export_thread.log.connect(self.log_panel.append_log)
        self.export_thread.finished.connect(
            lambda ok, path, _w=weights_path, _o=out_path, _e=expected_onnx:
            self._on_export_finished(ok, path, _w, _o, _e)
        )
        self.export_thread.start()
        self._append_management_log("模型管理", f"导出 ONNX: {Path(weights_path).name}", "已启动")

    def _on_export_finished(self, success: bool, onnx_path: str, weights: str, out_dir: str, expected_onnx: str):
        if success:
            path = onnx_path or expected_onnx
            file_size_mb = Path(path).stat().st_size / (1024 * 1024) if Path(path).exists() else 0
            self.log_panel.append_log(f"SUCCESS: ONNX 导出成功 -> {path} ({file_size_mb:.2f} MB)")
            self._append_management_log("模型管理", f"导出 ONNX: {Path(weights).name}", f"成功 → {path} ({file_size_mb:.2f} MB)")
            QMessageBox.information(
                self,
                "导出成功",
                f"ONNX 模型已成功导出:\n{path}\n\n文件大小: {file_size_mb:.2f} MB",
            )
        else:
            self.log_panel.append_log(f"FAIL: ONNX 导出失败，请查看上方日志获取详细错误信息")
            self._append_management_log("模型管理", f"导出 ONNX: {Path(weights).name}", "失败")
            onnx_files = list(Path(out_dir).rglob("*.onnx"))
            detail = "请在「运行日志」面板中查看详细错误信息。"
            if onnx_files:
                detail += f"\n\n不过，在输出目录中找到了 {len(onnx_files)} 个 ONNX 文件:\n" + "\n".join(str(f) for f in onnx_files[:5])
            QMessageBox.warning(self, "导出失败", detail)

    def on_quantize_onnx_requested(self):
        if not self._check_permission("can_switch_model", "量化 ONNX"):
            return

        try:
            self._do_quantize_onnx()
        except Exception as exc:
            import traceback
            self.log_panel.append_log(f"量化 ONNX 异常: {exc}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "量化失败", f"量化 ONNX 过程中发生异常:\n{exc}")

    def _do_quantize_onnx(self):

        try:
            model_path, _ = QFileDialog.getOpenFileName(self, "选择输入 ONNX 模型", "", "ONNX Model (*.onnx)")
        except Exception as exc:
            QMessageBox.critical(self, "文件选择失败", f"打开文件对话框异常: {exc}")
            return
        if not model_path:
            return

        try:
            out_path, _ = QFileDialog.getSaveFileName(self, "保存量化后模型为", str(Path(model_path).with_suffix(".quant.onnx")), "ONNX Model (*.onnx)")
        except Exception as exc:
            QMessageBox.critical(self, "保存路径选择失败", f"保存对话框异常: {exc}")
            return
        if not out_path:
            return

        already_fp16 = False
        try:
            import onnx
            model = onnx.load(model_path)
            already_fp16 = self._check_model_mostly_fp16(model)
        except ImportError:
            pass
        except Exception:
            pass

        if already_fp16:
            import shutil
            try:
                shutil.copy2(model_path, out_path)
            except Exception as exc:
                QMessageBox.critical(self, "复制失败", f"复制文件时出错:\n{exc}")
                return
            file_size_mb = Path(out_path).stat().st_size / (1024 * 1024)
            self.log_panel.append_log(f"模型已为 FP16，无需再次量化，已复制到: {out_path} ({file_size_mb:.2f} MB)")
            self._append_management_log("模型管理", f"量化 ONNX: {Path(model_path).name}", f"已为FP16，复制到 {Path(out_path).name}")
            QMessageBox.information(
                self,
                "量化完成",
                f"该模型已是 FP16 格式，无需再次量化。\n\n已复制到:\n{out_path}\n\n文件大小: {file_size_mb:.2f} MB",
            )
            return

        mode = QComboBox()
        mode.addItem("FP16", "fp16")
        mode.addItem("INT8", "int8")

        input_name = QLineEdit("images")
        calibration_dir = QLineEdit("")

        dialog = self._show_tool_dialog(
            "ONNX 量化选项",
            [
                ("mode", mode),
                ("input_name", input_name),
                ("calibration_dir", calibration_dir),
            ],
        )
        if dialog is None:
            return

        quant_mode = str(dialog.value("mode"))
        cmd = [
            sys.executable,
            str(Path(__file__).resolve().parents[1] / "tools" / "quantize_onnx.py"),
            "--input",
            model_path,
            "--output",
            out_path,
            "--mode",
            quant_mode,
            "--input-name",
            dialog.value("input_name") or "images",
        ]
        if quant_mode == "int8":
            calib_dir = dialog.value("calibration_dir")
            if not calib_dir:
                self.log_panel.append_log("取消 INT8 量化：未选择校准目录")
                return
            cmd.extend(["--calibration-data", calib_dir])

        self._start_tool_runner(cmd, cwd=str(Path(__file__).resolve().parents[1]))
        self._append_management_log("模型管理", f"量化 ONNX: {Path(model_path).name}", "已启动")

        quant_out_path = out_path
        runner = self.tool_runners[-1] if self.tool_runners else None
        if runner:
            runner.finished.connect(lambda rc, _out=quant_out_path:
                                    self._on_quantize_finished(rc, _out))

    @staticmethod
    def _check_model_mostly_fp16(model) -> bool:
        import numpy as np
        fp16_count = 0
        fp32_count = 0
        for init in model.graph.initializer:
            num_elems = max(1, int(np.prod(init.dims) if init.dims else 0))
            raw_len = len(init.raw_data) if init.raw_data else 0
            if raw_len == num_elems * 2:
                fp16_count += 1
            elif raw_len == num_elems * 4:
                fp32_count += 1
        total = fp16_count + fp32_count
        return total > 0 and fp16_count >= total * 0.8

    def _on_quantize_finished(self, return_code: int, out_path: str):
        if return_code == 0 and Path(out_path).exists():
            file_size_mb = Path(out_path).stat().st_size / (1024 * 1024)
            self.log_panel.append_log(f"SUCCESS: 量化完成 -> {out_path} ({file_size_mb:.2f} MB)")
            self._append_management_log("模型管理", f"量化 ONNX: {Path(out_path).name}", f"成功 ({file_size_mb:.2f} MB)")
            QMessageBox.information(
                self, "量化完成",
                f"FP16 量化成功！\n\n{out_path}\n\n文件大小: {file_size_mb:.2f} MB",
            )
        else:
            self.log_panel.append_log(f"FAIL: 量化失败，返回码={return_code}，请查看上方日志")
            self._append_management_log("模型管理", "量化 ONNX", "失败")
            QMessageBox.warning(self, "量化失败", f"量化过程出错 (返回码={return_code})。\n请在「运行日志」面板查看详细错误。")

    def on_benchmark_onnx_requested(self):
        if not self._check_permission("can_switch_model", "Benchmark ONNX"):
            return
        model_path, _ = QFileDialog.getOpenFileName(self, "选择 ONNX 模型", "", "ONNX Model (*.onnx)")
        if not model_path:
            return

        source = QLineEdit(self.control_panel.get_video_path() or "0")
        imgsz = QSpinBox()
        imgsz.setRange(32, 4096)
        imgsz.setSingleStep(32)
        imgsz.setValue(800)

        warmup = QSpinBox()
        warmup.setRange(0, 10000)
        warmup.setValue(10)

        limit = QSpinBox()
        limit.setRange(1, 100000)
        limit.setValue(300)

        providers = QLineEdit("")

        dialog = self._show_tool_dialog(
            "ONNX Benchmark 选项",
            [
                ("source", source),
                ("imgsz", imgsz),
                ("warmup", warmup),
                ("limit", limit),
                ("providers", providers),
            ],
        )
        if dialog is None:
            return

        cmd = [
            sys.executable,
            str(Path(__file__).resolve().parents[1] / "tools" / "benchmark_onnx.py"),
            "--model",
            model_path,
            "--source",
            dialog.value("source") or "0",
            "--imgsz",
            str(dialog.value("imgsz")),
            "--warmup",
            str(dialog.value("warmup")),
            "--limit",
            str(dialog.value("limit")),
        ]
        provider_text = dialog.value("providers")
        if provider_text:
            cmd.extend(["--providers", provider_text])

        self._start_tool_runner(cmd, cwd=str(Path(__file__).resolve().parents[1]))
        self._append_management_benchmark_record(
            model_name=Path(model_path).stem,
            input_size=str(dialog.value("imgsz")),
            device=self._device_label(),
            note="已启动 Benchmark",
        )
        self._append_management_log("模型管理", f"Benchmark ONNX: {Path(model_path).name}", "已启动")

    def closeEvent(self, event):
        try:
            if getattr(self, "tray_icon", None) is not None:
                self.tray_icon.hide()
        except Exception:
            pass

        try:
            if self.worker.isRunning():
                self.worker.stop()
                self.worker.wait(3000)
                if self.worker.isRunning():
                    self.worker.terminate()
                    self.worker.wait(2000)
        except Exception:
            pass

        for t in [self.export_thread, self._model_eval_thread]:
            if t is not None and t.isRunning():
                try:
                    t.quit()
                    t.wait(3000)
                    if t.isRunning():
                        t.terminate()
                        t.wait(2000)
                except Exception:
                    pass

        for runner in list(self.tool_runners):
            if runner.isRunning():
                try:
                    runner.quit()
                    runner.wait(2000)
                    if runner.isRunning():
                        runner.terminate()
                        runner.wait(1000)
                except Exception:
                    pass
            self.tool_runners.remove(runner)

        try:
            self.history_db.close()
        except Exception:
            pass
        try:
            self.auth_manager.close()
        except Exception:
            pass
        super().closeEvent(event)
